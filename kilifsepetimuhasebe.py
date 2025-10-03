from __future__ import annotations

# Embedded Assets
LOGO_BASE64 = """iVBORw0KGgoAAAANSUhEUgAAAGQAAAAaCAYAAABByvnlAAAACXBIWXMAAC4jAAAuIwF4pT92AAAKamlUWHRYTUw6Y29tLmFkb2JlLnhtcAAAAAAAPD94cGFja2V0IGJlZ2luPSLvu78iIGlkPSJXNU0wTXBDZWhpSHpyZVN6TlRjemtjOWQiPz4gPHg6eG1wbWV0YSB4bWxuczp4PSJhZG9iZTpuczptZXRhLyIgeDp4bXB0az0iQWRvYmUgWE1QIENvcmUgOS4xLWMwMDIgNzkuYTZhNjM5NiwgMjAyNC8wMy8xMi0wNzo0ODoyMyAgICAgICAgIj4gPHJkZjpSREYgeG1sbnM6cmRmPSJodHRwOi8vd3d3LnczLm9yZy8xOTk5LzAyLzIyLXJkZi1zeW50YXgtbnMjIj4gPHJkZjpEZXNjcmlwdGlvbiByZGY6YWJvdXQ9IiIgeG1sbnM6eG1wPSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvIiB4bWxuczpkYz0iaHR0cDovL3B1cmwub3JnL2RjL2VsZW1lbnRzLzEuMS8iIHhtbG5zOnBob3Rvc2hvcD0iaHR0cDovL25zLmFkb2JlLmNvbS9waG90b3Nob3AvMS4wLyIgeG1sbnM6eG1wTU09Imh0dHA6Ly9ucy5hZG9iZS5jb20veGFwLzEuMC9tbS8iIHhtbG5zOnN0RXZ0PSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvc1R5cGUvUmVzb3VyY2VFdmVudCMiIHhtbG5zOnN0UmVmPSJodHRwOi8vbnMuYWRvYmUuY29tL3hhcC8xLjAvc1R5cGUvUmVzb3VyY2VSZWYjIiB4bXA6Q3JlYXRvclRvb2w9IkFkb2JlIFBob3Rvc2hvcCBDQyAyMDE5IChXaW5kb3dzKSIgeG1wOkNyZWF0ZURhdGU9IjIwMjEtMDgtMTdUMjA6MDg6MzIrMDM6MDAiIHhtcDpNb2RpZnlEYXRlPSIyMDI1LTA5LTAyVDE1OjA1OjI1KzAzOjAwIiB4bXA6TWV0YWRhdGFEYXRlPSIyMDI1LTA5LTAyVDE1OjA1OjI1KzAzOjAwIiBkYzpmb3JtYXQ9ImltYWdlL3BuZyIgcGhvdG9zaG9wOkNvbG9yTW9kZT0iMyIgeG1wTU06SW5zdGFuY2VJRD0ieG1wLmlpZDo1MGFlNzI3OC1jYjNjLTQ0NDQtYTAxYi1hYzliMDY1MjBlZTEiIHhtcE1NOkRvY3VtZW50SUQ9ImFkb2JlOmRvY2lkOnBob3Rvc2hvcDpjOWQ0MWZmOC1lYjBiLWNlNDktYTk2Yy0zY2RkYTE2MDI4MDMiIHhtcE1NOk9yaWdpbmFsRG9jdW1lbnRJRD0ieG1wLmRpZDoxMzU5YzgzMy0zODVlLWEwNGItYTgwNi1iZmZmMDc1NDcwOWUiPiA8eG1wTU06SGlzdG9yeT4gPHJkZjpTZXE+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJjcmVhdGVkIiBzdEV2dDppbnN0YW5jZUlEPSJ4bXAuaWlkOjEzNTljODMzLTM4NWUtYTA0Yi1hODA2LWJmZmYwNzU0NzA5ZSIgc3RFdnQ6d2hlbj0iMjAyMS0wOC0xN1QyMDowODozMiswMzowMCIgc3RFdnQ6c29mdHdhcmVBZ2VudD0iQWRvYmUgUGhvdG9zaG9wIENDIDIwMTkgKFdpbmRvd3MpIi8+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJjb252ZXJ0ZWQiIHN0RXZ0OnBhcmFtZXRlcnM9ImZyb20gaW1hZ2UvcG5nIHRvIGFwcGxpY2F0aW9uL3BkZiIvPiA8cmRmOmxpIHN0RXZ0OmFjdGlvbj0ic2F2ZWQiIHN0RXZ0Omluc3RhbmNlSUQ9InhtcC5paWQ6M2M4NjU3MzctMGQwNS0xODQ0LWE5MzEtZWNlNGNjMGQzMDdjIiBzdEV2dDp3aGVuPSIyMDIxLTA4LTE3VDIwOjI2OjQ0KzAzOjAwIiBzdEV2dDpzb2Z0d2FyZUFnZW50PSJBZG9iZSBQaG90b3Nob3AgQ0MgMjAxOSAoV2luZG93cykiIHN0RXZ0OmNoYW5nZWQ9Ii8iLz4gPHJkZjpsaSBzdEV2dDphY3Rpb249InNhdmVkIiBzdEV2dDppbnN0YW5jZUlEPSJ4bXAuaWlkOmViNTcxZDc3LWJhNTctZDg0OS04Y2NiLTk2M2NhNjM5ZDgwNSIgc3RFdnQ6d2hlbj0iMjAyMS0wOC0xN1QyMDoyOTo0NiswMzowMCIgc3RFdnQ6c29mdHdhcmVBZ2VudD0iQWRvYmUgUGhvdG9zaG9wIENDIDIwMTkgKFdpbmRvd3MpIiBzdEV2dDpjaGFuZ2VkPSIvIi8+IDxyZGY6bGkgc3RFdnQ6YWN0aW9uPSJjb252ZXJ0ZWQiIHN0RXZ0OnBhcmFtZXRlcnM9ImZyb20gYXBwbGljYXRpb24vcGRmIHRvIGltYWdlL3BuZyIvPiA8cmRmOmxpIHN0RXZ0OmFjdGlvbj0iZGVyaXZlZCIgc3RFdnQ6cGFyYW1ldGVycz0iY29udmVydGVkIGZyb20gYXBwbGljYXRpb24vcGRmIHRvIGltYWdlL3BuZyIvPiA8cmRmOmxpIHN0RXZ0OmFjdGlvbj0ic2F2ZWQiIHN0RXZ0Omluc3RhbmNlSUQ9InhtcC5paWQ6MTUwYjJkNjEtZDUwZS03ZDRlLTk3MjUtMmI1ZmU3YjExOThmIiBzdEV2dDp3aGVuPSIyMDIxLTA4LTE3VDIwOjI5OjQ2KzAzOjAwIiBzdEV2dDpzb2Z0d2FyZUFnZW50PSJBZG9iZSBQaG90b3Nob3AgQ0MgMjAxOSAoV2luZG93cykiIHN0RXZ0OmNoYW5nZWQ9Ii8iLz4gPHJkZjpsaSBzdEV2dDphY3Rpb249InNhdmVkIiBzdEV2dDppbnN0YW5jZUlEPSJ4bXAuaWlkOjUwYWU3Mjc4LWNiM2MtNDQ0NC1hMDFiLWFjOWIwNjUyMGVlMSIgc3RFdnQ6d2hlbj0iMjAyNS0wOS0wMlQxNTowNToyNSswMzowMCIgc3RFdnQ6c29mdHdhcmVBZ2VudD0iQWRvYmUgUGhvdG9zaG9wIDI1LjExIChXaW5kb3dzKSIgc3RFdnQ6Y2hhbmdlZD0iLyIvPiA8L3JkZjpTZXE+IDwveG1wTU06SGlzdG9yeT4gPHhtcE1NOkRlcml2ZWRGcm9tIHN0UmVmOmluc3RhbmNlSUQ9InhtcC5paWQ6ZWI1NzFkNzctYmE1Ny1kODQ5LThjY2ItOTYzY2E2MzlkODA1IiBzdFJlZjpkb2N1bWVudElEPSJ4bXAuZGlkOjEzNTljODMzLTM4NWUtYTA0Yi1hODA2LWJmZmYwNzU0NzA5ZSIgc3RSZWY6b3JpZ2luYWxEb2N1bWVudElEPSJ4bXAuZGlkOjEzNTljODMzLTM4NWUtYTA0Yi1hODA2LWJmZmYwNzU0NzA5ZSIvPiA8L3JkZjpEZXNjcmlwdGlvbj4gPC9yZGY6UkRGPiA8L3g6eG1wbWV0YT4gPD94cGFja2V0IGVuZD0iciI/Pp/Hqx4AABOSSURBVGiB7ZpplFXFtcd/VWe6871Nd0OL0NBM2gIyqiBRMYgi8qKQ4Ul44tL4UPMwicQhGpOQBFEJccqwxCFi5AlGjYqiqICIMguNMggiKAhN09jN7eGOZ6j34dy+dAdwJVl5a/khe62z7jm7du2qW/+q+u9d5wilFP+Wr47oAG//OUaXMujWVfvlsrWBq5tbtE669tVFSkpobBa0phWVFQqlBEoBApSC1pRC0yAUFAAoD1rTHXV5W+E6vj/TBCkF2awCASiwLIHngespDF2QyykCliDvKHQNNE2Qt8F1FIGAQAhwXcjnfR+m4dsoBdmcgkL/LFMghBK27Xqn9Q5vHzVx9EzVcupbouTXQAGQshJOd135vf+4Qd6yan0r4Pm1v3LiFX4F484XXDgyxJQZOcAGZKFMARp+/9v0f6sTBb0oXG6hrlawbWtLtPMpC3Zt/lQ7H2472/Y6F3+I24+l1+6+/tzZtx19847Z3x4K1LTV5nCDmL7odXHLqvWKSFQnFjeJxQziCQvDcIDUcVc0qhMv2LW/EgmLaFQ/zl7T8iQSVtFvLGZ0tBFZvzxukkhYHcpM00XXbSCLpjlAjquvMLhthgkigGkaRKMGQhhcf6XJR5sM9q6LM/b8EkBy01STz7abfPh2nJHDEwgh+dXNJls2BNm2Kcz4r8fo3VPjw3cDbFocYOvaCIP6R/j2eMm8WXGGDDDZ9FaMM04LM39OiHGjI4DGb27X2Lw1Tq+qKEIIxowy2fR2gO1rAzwyuxNSD1PRWbLuOYsPV1p88HqMAdVhEJJ4wgBKuHNOkvqt+29sg0gHsB0Rb037M8HQ/WUvNUljYyMA48aNY8CAAViWRX19PatXr2bHjh2Ew2Esy8LzjqHuOHkAxowZg2VZRf0HH3xAU1MTuq5j2zksy+KKK64gn/fta2pqaGlpwTAMUqlWBgwYQGVlJYZhsGXLFoQQXH755aTTaV58+T12fbKL5oNhhO5vRXkHhPR45PEAv75L0PWUJNO+E+etVYKHZxvc9Ru4eFQT836Z4MzLdKZMALdZMv76Vhxb5/BhyZ/mZ3jgoSA//5nL3n05xo7w+PoIeH+bZOhFeUYN1RhxpsOTL+qAYvz5gjMGKEYN1dj7qWDQaQ5De1lcelWW15dKnl8aYn3NF5xzSZBrpnqs2ZIk45hEwhoCCIQUThr21nY7rfPgdoBISdY0cAFNKZBSkkwm6dq1K8899xznnnsu7cV1XWbNmsXMmTMRQqDrerGstbWVXr16sWzZsg51br75Zh588EESiQSe59Hc3Mx9991Hv379AHjggQeYMWMGpmmSz+dZsmQJlZWVAIwdOxYpJZMnT2bTpk2sWFNPp9Ij2I5d9G/okMsI3luc57rr4vQaBFn3KKWlGl6LxstLJQGV4ZvjHUCwr1YR7Szp1ilMzcd5MjmP516FB27TeOFVaGl12b1P0Jj0OL2XDs0wtFqx94Bg/QcuFZ0Vh76w6PSeYvRZiqdfkLSkPI4ckUTLLA5stKk77BKwNNwDir7DTXbth41bbaIRf9IrBboGkWg43vY/JCeQbDaL53k8//zzHcA4evQoAJqm8Ytf/IJJkybR2tqKEB35ZsCAAcf5PPPMM4v3gUAAx3GYM2dOUTd58mQ0TSOfz3PBBRcUwaipqWH9+vVUVVWxYcMGPv30UxKJEkpLo3jusZVpGRAKSs77rsP2XY3UfhIiFovQ2uqStxXlpYpIVJDJ+vb1XyhOH5bn7h8LSkt87uhSDqDoXOYz8J7PwdAV1X0sdr0r+NqZkM3rZNOKMSMEyaTGwwvynD1QADrJFkG0LM/vZ9vkXIuP9kDAFHiey+Xjba64SEN5Atc9Fi+5LhiGlv9SQNLpNEOGDGHkyJHF53HjxtGtWzfuu+++ot20adMKTn1SbAOmf//+x/k844wzALBtG6UUuq6zaNEimpqaAKioqGDMmDEAXHnllcV6jz32GC0tLdTX17N7924ikQgBI89HH32GaZpFu0wOQmGP6VdFGPtNk7q9R1k8L0AuJwkkBPv3C5KNirISv4/9eklWPW8y6soURxryx4ai3dz69KAgFrYZOQj+stxjwAgXpQTgUVEuuXiiy1VTPMpKADRiIZfkAZPBX4Mep7iM+ZpJS9rBCFicNVxy+5wkgaCLYRwbdiGgfepxQkAAqqqqivcHDhzgjTfeIJ1O88ILLxw3yI7jx49tXNK2Gj755BN27txZBKmsrIx0Og1AJBIhlUrx9NNPF/1deumlAEycONEf5EyGhQsXIqXk5Zdf5r333mPx4sWsXrmQ0k6RDv31FDQ2Kn73M491y4OUxAUHD2UBwf69Ns/Mhxk/Mlm9yQE8KrtAZXcJyOJEMg2gpyBg+c+pVkHnTnkSAx2WrxPQxUHX/Ajq2m8a7P0Y7n4wR8UgOLWrTj7vUdFLkrI19FKP0rginRHQ1eaNFwXzf5Og+ykWzc0uJxP9ZAV1dXXF+8rKSvr27cvu3btpbGzkxhtvREpJXV0doVCoyCHZrL8fDBw4EICVK1eSSqU4/fTTiUQiVFdX8+677wLHVtMTTzzB9OnTATj77LO55JJL6NKlCwCLFi0imUwSj8cRQlBTU1PoURxN69jfoAVCSXqPzXHV5By5jOCRp3KApGpMjt/NdfjrK5J5T+UAuHyaixlOEQ75eU04Ajs/lUy8JMfSlR5SA9OQjJvmMWJwlg0fwMSLYfNHKYJBxf3zFS+/afNFUiObznCwVrJoqaT8ZxmamxTjJrnUJz2kZjLx8ixjLhR0jgVRSiG+JKM4ISCBQKC4X1dVVREIBIrkvmfPHvbs2VO0LS0txfM8lFLkcjkqKyuLRL1t2zYaGhqKtv379y8CopQiFAqxZcsW3n//fYYPH86wYcM6rJjHH38c8IMMpRSJRAKAZBKUah/PFwhSh737YeVqkw0bIZPLMm60QTod5qabbbp31Ugm81z7HQ1RorP0Fcg7OUxLIIWiNKGhSuF/rrOYv8glHMwjZJD7n7AJhhyiXQwuiFo8szjHS8s9elYaYChMU+PGGzyCTpT7H89xwTkeeRGgZluOslLF/rogN93pAmmCIUk0qnXgkfZywi0rFArhOA5XXXVVUTdo0CBqamqKhB2NRikpKcF13Q57YHV1NYZhALBnzx527NhRLGtbOe2BB3j00UcBsCyL8vJyAHbs2MGaNWsIBAL8vcc7LS1w7lBYuUpj7p0BUIIrx9m885bJw78MMPcOf1k9dIfONRMVbz9l0rWLRXOLS0uL4LpveTw60+CG7wp+9v0gLa0uy+frmEGTObdIZnxPcuu1MHpEgO9c4vDMXIumpOQbFzr88TcGp5T7af6sHypWrAoQiQVobrVZu1Dn0q/7KYBlipOCAScBxPM84vE4q1evLu7nAP369WPr1q1MmTKFlpYWHMc5LsJqP+gHDx5k//79x5W15xxN03j22WdpbW3t4KcNpFAo9OUodBDFWQN1crsUuvSz5I/2Skh7fG2oxtbdHqD49KDGjB9Dc0uGPj1l8QglEoKnn9KZ+1CWswe6KHQO1HmgPHpXarz2ksWsP2RJNtl4SpDJetg23DrHZucqjVvvcTDNPIcbAhxZ73HRSMjn4MBhj3ze+7KOF+WkpC6lJBaL8dJLL3HOOeewb9++YtmCBQuYNm0aLS0tHZJC6Bjy1tfX09DQQCqVAvwtKxaLFYkdfHJvbm7mT3/6U1HX2NjIggULkFIe5/9kUji1YmA/yYYtcEq5B2housZ7KxSD+uUoifkJHSjiCUUmK7GdY7M1b0MwpAgGBc2tAikVnudHQkcaoaKLT+YlcZ3WlCJfSIMquwp0SxEMC6r7CPKeRs1uj7PP9HvledDS6rf7ZfwBJwFECIFt22QyGXRdZ8OGDfTv35933nmnaDNv3jyGDBlCc3MzQghyOZ8s21bBxx9/zMGDB1FKsW7dOgA6depEdXV1h0GW0u9Cmw3A5s2baWhoIBwOf3nv20nOBk2DUUMUssxl/PmCTp10upR6bNwBBw7lGDnYb6uiTDB0qMbpvQSFbgMQsKBPH8nw4RqhoMJ2FF07C1xP0L1C0ZCRZB2Pfr38QKCizPdnGL5dJiUYdoZk7GhFeVWWYWdogM4p5dC/v0FZJwP35AGWPx4nUiql0DQN0zSxLIt4PE4qlWL06NFs3769aHfHHXeglMLzPDKZDBUVFcVQ2DRNfvKTn3DnnXcWoyY4Pkdp44dgMFjUaX8bQv0d4jjQ/VRYuwXGT7L531fyVHV3Wf+Bx6btLnPnC1ZtzAOC9R/muHySy6N/cVi/OYfQBLoOK9YpZDxD797wk99m0KRixToXJ+dx670254zIEY0IHluUpe6I4NW382ia4vAXgldW2GiaR8CS3P2gwzX/nWV/rY0VMHh9VY7xE3JcOcEgk/EDkJPJCaOs5uZmLrvsMu6//34AlixZwi233ILjONx7773FSKhte7Jtf+1WV1cXibpnz57cc889x/luqyOE6JgQSXnc/T/yriYYgJaU5OZ7PdJZnatvs6nootizTyeXz5JJQyzqUFpq8M2bbKJhaGjU6HOaxsEDGnbeZtk6jfVbPQ4cdBGaRu9Kk2/ckKN7pUZd2uL7N3vU1jtUnqqxda/JZ087JBKKQ0cMrr3LIRr1+OtynbrDLgiD23+bJxxSfPsHHuefpzj4uY5tK6QUJwXlpKReUlJCnz596NOnDwMHDiwScX19fdGujdDbBq797K+trWXjxo1s3LiRzz777DhA2vz9q8RT0NCgeO4BxWcrQoDB2oWKv/4+RCatgVDsWQmvPBIin9NpaIS1z+q8tEBjxZ+DdC4L0rtbnvmzw1T1MNjzloHn6Qztr9jwqsGixzVGDbNwXMnCBySbXtF4ck4Ax9G46wZ4YlaY5mZ463HF/Dlhrpmk88efBzl6FH79I8lDczTenB/i66NC/1xiuHnz5uL90KFDmTJlCnV1dcycObOo37VrFwCGYZDJZDqcV11//fW8+uqrxfqbNm0CfI4JhUKk02mi0eg/M/YnlLYZV1HqcOpgj1FnGfQc0koq67/XGHKGoqy7wnPAChrkczkGVUsm3eAx754Uky4OsWSFR1U3l4X3mqypybF3n8tt13kk0DjvIhsrkCKX1bjt3gxvPxNm6o8cmppyDDpNo1OJh+cJIsE8V38rjCF1QgEHpeCum0z+8weKnhVJhvUP8c7ajszePlI94QqJRCJs27aNRYsWAZBIJFiwYAHLli0rnm8B3H333QghirP9wgsvLJa1HZmAT/BtRF5RUcHgwYOL519t0v6oPh6P88+JYsdek8adBt+6RHB4p8XHn/kll5wnWbzIYM1Gh/OGS5QHBw65/PxuQT5v8Nc38oBBr35ZzrlS8sCfLcDhYL0GMZuXnjU5Z3AQ8Kj7QnK0QVJ7pHDEkhE0tQB4rPswQG2t4LIxgo/3+XnJ9p0wpNogVm6yusbFMI8B4L9p9Iy25xMC0kbmU6dO5Q9/+EOHbBtgw4YNTJgwgY0bNxKLxchms5SXlyOlpKmpia1bt3Lo0CF0XSccDtPa2srSpUtJJpM0NzfTo0ePDrNCCEFzczPJZJJkMsnnn39+XH7z90omL/lkn8Ol50n27JO4jg98726CWKWksp9Hnx6+b02D+loXhaSsk44QHl8cEhz9NMdN/6UDBuEQpJOCJ+dn2bXXDwpCAX8gQ4HCNGjHB5ap8coKm3ifLOGQH5wYBuza43LtRJg8wSKbPRZl6jpksnZj8bkwIEHbQfPvfQ6JRCIcPXqU6dOnM3v2bPr27YtlWdTV1fHhhx8C/soRQhCJ+Ad9F1xwAZ7nkc1mEUIUtyRN05gyZUoxyVNKEYvFip2KxWIsWbKE5cuXAz6/xGKxfwiUtjHpUylJpR1O7+mx/5Ckd6U/S0cOlhw1FN17OIwZYfLIAkGvvgaTfwhPzsnyjTEhnnvNQ3khvn9jmoVPuUz/VQjPTdKpPMpH+zxKE1kO13vomiLRTVA4kKCiDEoS/tweMUhw61y4Zq/NyMF+2N6rm6Jzpc7nnzu4zrFXw0JAKg+Gu7tIzD4geEdKE/674EzOf2voui7hSBzP86itraW2trb45w0jQDAYLBwBKITQSGfyHDlypGAhicbiOC6g/PKmphaSySQAphUiYFl4nip0TJJKZYtvKBE6sWi0WH4iKU1AvFKibIENmNL/h+9tcqmth9feUaSz0PNUiZCKrbs1pv3UZfTwHFdcqoPQefZFGwzBa6tg6y6H1rTGyg02L6yEZxfl6N/P4K21GsvXpJl3v8a7S01+vjtHQ1Kw+EWHo0m/zTfXKCLhPAjBa6scNu8Q/PQhSVX3PAjJ1NvzzLhdcWgnzHksg6YLPAXZtA/BKeVHWtr+l1BK8cSsUFVZqTb1+besmQteLCynr+I3Dm2iYMKFHucOs7hzbuFjAiH8Pp8IRI1j3zH4RqBJ6HCm1PbRQtsMVggNlNumK3w2ImXhO4WCnRSFNlWhTHVsp6gr1AeELlAOCOHy8EyD6bd2609wyw4oADLjmiAlcRhSrc840qRf7dh0/hcN3f+LaBLqG6GpxaNvD4FX+AxIALbbcS4pwNDAdvz92vP8y9B9naZRPB6R0tcZup9oSlnQFY5IpPRfuSrlt6NrPh6O27GdNp9K+b4UkM35etPwy2wbdVqVu/miCeFZpMLr6PEJUADk3/LVkf8D7HxlvR8WFscAAAAASUVORK5CYII"""


# -*- coding: utf-8 -*-
"""
NES Toolkit + Kılıf Sepeti Envanter – Tek Dosya / Tek Program (v3.1) + UI rev.
------------------------------------------------------------------------------
• Tek ana Tk uygulaması (tek root), 4 sekme:  [Fatura], [Rapor],
  [PDF Eşleştirici], [Envanter]  (+ Hakkımızda)
• Tema: Windows 11 esintili minimalist koyu/açık (varsayılan koyu).
• UI/UX: Kart paneller, tutarlı tipografi, marka yeşili aksan.
• Fonksiyonel kısım korunmuştur; yalnızca görünüm ve “Hakkımızda” eklendi.

Not: PyInstaller önerisi (konsolsuz):
pyinstaller --noconsole --windowed --name "NES_Toolkit" --icon "app.ico" envanter_gui.py
"""

# ======= Standart Kütüphane =======
import os
import re
import sys
import time
import math
import json
import zipfile
import calendar
import threading
import webbrowser
import ctypes
import traceback
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional

# ======= 3rd Party =======
try:
    import requests
except Exception:
    requests = None  # Ortamda yoksa GUI’de uyarı verilecek.

import pandas as pd
import numpy as np
from PyPDF2 import PdfReader
import openpyxl

# ======= Tk =======
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter.font import Font

# Win registry (opsiyonel)
try:
    import winreg
    HAS_WINREG = True
except Exception:
    HAS_WINREG = False

# Opsiyonel: tkcalendar / PIL (yoksa devam)
try:
    from tkcalendar import DateEntry as _DateEntry  # noqa
    HAS_CALENDAR = True
except Exception:
    HAS_CALENDAR = False

try:
    from PIL import Image, ImageTk  # noqa
    import base64  # noqa
    HAS_PIL = True
except Exception:
    HAS_PIL = False


# ===========================================================================
# UI Teması & Genel Ayarlar (YENİ)
# ===========================================================================

# Modern Brand Colors
BRAND_GREEN = '#00D4AA'   # Modern teal-green
BRAND_BLUE = '#0066FF'    # Vibrant blue
BRAND_PURPLE = '#8B5CF6'  # Modern purple

LIGHT_COLORS = {
    'primary'   : '#1A1A1A',
    'secondary' : BRAND_BLUE,
    'accent'    : BRAND_GREEN,
    'accent_hover': '#00B894',
    'warning'   : '#FF6B6B',
    'success'   : '#51CF66',
    'background': '#F8FAFC',
    'card'      : '#FFFFFF',
    'card_hover': '#F1F5F9',
    'text'      : '#1A1A1A',
    'text_light': '#64748B',
    'text_muted': '#94A3B8',
    'border'    : '#E2E8F0',
    'border_light': '#F1F5F9',
    'hover'     : '#F8FAFC',
    'shadow'    : 'rgba(0, 0, 0, 0.1)',
    'gradient_start': '#FFFFFF',
    'gradient_end': '#F8FAFC'
}

DARK_COLORS = {
    'primary'   : '#F1F5F9',
    'secondary' : '#3B82F6',
    'accent'    : BRAND_GREEN,
    'accent_hover': '#00B894',
    'warning'   : '#F87171',
    'success'   : '#34D399',
    'background': '#0F172A',
    'card'      : '#1E293B',
    'card_hover': '#334155',
    'text'      : '#F1F5F9',
    'text_light': '#CBD5E1',
    'text_muted': '#94A3B8',
    'border'    : '#334155',
    'border_light': '#475569',
    'hover'     : '#1E293B',
    'shadow'    : 'rgba(0, 0, 0, 0.3)',
    'gradient_start': '#1E293B',
    'gradient_end': '#0F172A'
}

COLORS = LIGHT_COLORS.copy()
PADDING = {'xs': 4, 'small': 8, 'medium': 16, 'large': 24, 'xlarge': 32, 'xxl': 48}
BORDER_RADIUS = {'small': 6, 'medium': 12, 'large': 16, 'xlarge': 20}

SETTINGS_FILE = "settings.json"
APIS_FILE     = "apis.json"
REG_PATH      = r"Software\NESFaturaIndirici\APIs"

HEADERS_TEMPLATE = {
    "Authorization": "",
    "Accept": "application/json",
    "Content-Type": "application/json"
}

APP_SETTINGS = {
    "theme": "dark",                  # varsayılan koyu
    "window_size": "1280x860",
    "last_tab": 0,
    "last_api": "",
    "default_save_folder": "",
    "auto_zip": True,
    "auto_open_folder": True,
    "invoice_folder": ""
}


def apply_modern_theme(root, existing_style=None):
    """
    Ultra-modern theme with rounded corners, gradients, and contemporary styling.
    """
    style = existing_style or ttk.Style(root)
    try:
        style.theme_use('clam')
    except Exception:
        pass

    # Modern typography with better readability
    base_font = ('Segoe UI', 11)  # Increased base size for better readability
    medium_font = ('Segoe UI', 12)
    title_font = ('Segoe UI', 13, 'bold')
    large_title_font = ('Segoe UI', 15, 'bold')
    button_font = ('Segoe UI', 11, 'bold')  # Bold for buttons

    # Modern Frames with subtle gradients
    style.configure('TFrame', background=COLORS['background'], relief='flat', borderwidth=0)
    style.configure('Card.TFrame', 
        background=COLORS['card'], 
        relief='flat', 
        borderwidth=1,
        bordercolor=COLORS['border'])
    style.configure('Elevated.TFrame', 
        background=COLORS['card'], 
        relief='solid', 
        borderwidth=1,
        bordercolor=COLORS['border_light'])
    
    # Modern LabelFrames with subtle gradients
    style.configure('TLabelframe', 
        background=COLORS['card'], 
        bordercolor=COLORS['border'],
        borderwidth=1,
        relief='solid')
    style.configure('TLabelframe.Label', 
        background=COLORS['card'], 
        foreground=COLORS['accent'], 
        font=('Segoe UI', 12, 'bold'),
        padding=(8, 4))

    # Enhanced Labels
    style.configure('TLabel', 
        background=COLORS['background'], 
        foreground=COLORS['text'], 
        font=base_font)
    style.configure('Card.TLabel', 
        background=COLORS['card'], 
        foreground=COLORS['text'], 
        font=base_font)
    style.configure('Title.TLabel', 
        background=COLORS['card'], 
        foreground=COLORS['accent'], 
        font=title_font)
    style.configure('LargeTitle.TLabel', 
        background=COLORS['card'], 
        foreground=COLORS['primary'], 
        font=large_title_font)

    # Modern Entry & Combobox with enhanced focus states
    style.configure('TEntry',
        fieldbackground=COLORS['card'], 
        foreground=COLORS['text'],
        bordercolor=COLORS['border'], 
        lightcolor=COLORS['border_light'], 
        darkcolor=COLORS['border'],
        insertcolor=COLORS['accent'], 
        relief='solid', 
        borderwidth=1,
        padding=14, 
        font=base_font
    )
    style.map('TEntry',
        fieldbackground=[('focus', COLORS['card'])],
        bordercolor=[('focus', COLORS['accent'])],
        lightcolor=[('focus', COLORS['accent'])],
        darkcolor=[('focus', COLORS['accent'])]
    )
    
    style.configure('TCombobox',
        fieldbackground=COLORS['card'], 
        background=COLORS['card'], 
        foreground=COLORS['text'],
        arrowcolor=COLORS['text_muted'], 
        bordercolor=COLORS['border'], 
        relief='solid',
        borderwidth=1,
        padding=12, 
        font=base_font
    )
    style.map('TCombobox',
        fieldbackground=[
            ('readonly', COLORS['card']), 
            ('focus', COLORS['card']),
            ('hover', COLORS['card_hover'])
        ],
        bordercolor=[
            ('focus', COLORS['accent']),
            ('hover', COLORS['border'])
        ],
        arrowcolor=[
            ('focus', COLORS['accent']),
            ('hover', COLORS['secondary'])
        ]
    )
    
    # Modern dropdown styling
    root.option_add('*TCombobox*Listbox.background', COLORS['card'])
    root.option_add('*TCombobox*Listbox.foreground', COLORS['text'])
    root.option_add('*TCombobox*Listbox.selectBackground', COLORS['accent'])
    root.option_add('*TCombobox*Listbox.selectForeground', '#ffffff')
    root.option_add('*TCombobox*Listbox.borderWidth', '2')
    root.option_add('*TCombobox*Listbox.relief', 'solid')

    # Modern Notebook / Tabs with subtle elevation
    style.configure('TNotebook', 
        background=COLORS['background'], 
        borderwidth=0, 
        tabmargins=[4, 6, 4, 0])
    style.configure('TNotebook.Tab',
        background=COLORS['border_light'], 
        foreground=COLORS['text_muted'],
        padding=[20, 12], 
        font=('Segoe UI', 11, 'bold'),
        borderwidth=0,
        relief='flat'
    )
    style.map('TNotebook.Tab',
        background=[
            ('selected', COLORS['card']), 
            ('active', COLORS['card_hover'])
        ],
        foreground=[
            ('selected', COLORS['accent']), 
            ('active', COLORS['text'])
        ]
    )

    # Modern Treeview with enhanced visual hierarchy
    style.configure('Treeview',
        background=COLORS['card'], 
        fieldbackground=COLORS['card'],
        foreground=COLORS['text'], 
        bordercolor=COLORS['border'],
        borderwidth=1,
        relief='solid',
        rowheight=36, 
        font=base_font)
    style.configure('Treeview.Heading',
        background=COLORS['background'], 
        foreground=COLORS['text'],
        relief='flat',
        borderwidth=0,
        font=('Segoe UI', 11, 'bold'))
    style.map('Treeview',
        background=[
            ('selected', COLORS['accent']),
            ('focus', COLORS['card_hover'])
        ],
        foreground=[
            ('selected', '#ffffff'),
            ('focus', COLORS['text'])
        ]
    )

    # Standard buttons with subtle elevation effect
    style.configure('TButton', 
        background=COLORS['card'], 
        foreground=COLORS['text'], 
        borderwidth=1, 
        relief='solid', 
        bordercolor=COLORS['border'],
        padding=(16, 14), 
        font=button_font
    )
    style.map('TButton',
        background=[
            ('pressed', COLORS['border_light']), 
            ('active', COLORS['card_hover']),
            ('disabled', COLORS['background'])
        ],
        bordercolor=[
            ('pressed', COLORS['border']), 
            ('active', COLORS['accent']),
            ('disabled', COLORS['border_light'])
        ],
        foreground=[
            ('active', COLORS['accent']),
            ('disabled', COLORS['text_muted'])
        ]
    )

    # Ultra-modern buttons with subtle shadows and enhanced hover states
    style.configure('Accent.TButton', 
        background=COLORS['accent'], 
        foreground='#ffffff', 
        borderwidth=0, 
        relief='flat',
        padding=(20, 18), 
        font=button_font
    )
    style.map('Accent.TButton',
        background=[
            ('pressed', COLORS['accent_hover']), 
            ('active', COLORS['accent']),
            ('disabled', COLORS['text_muted'])
        ],
        relief=[('pressed', 'flat'), ('active', 'flat')]
    )

    # Secondary Button
    style.configure('Secondary.TButton', 
        background=COLORS['background'], 
        foreground=COLORS['text'], 
        borderwidth=1, 
        relief='solid', 
        bordercolor=COLORS['border'],
        padding=(18, 16), 
        font=button_font
    )
    style.map('Secondary.TButton',
        background=[
            ('pressed', COLORS['border_light']), 
            ('active', COLORS['card']),
            ('disabled', COLORS['background'])
        ],
        bordercolor=[
            ('pressed', COLORS['border']), 
            ('active', COLORS['accent']),
            ('disabled', COLORS['border_light'])
        ],
        foreground=[
            ('active', COLORS['accent']),
            ('disabled', COLORS['text_muted'])
        ]
    )

    # Warning Button
    style.configure('Warning.TButton', 
        background=COLORS['warning'], 
        foreground='#ffffff', 
        borderwidth=0, 
        relief='flat',
        padding=(18, 16), 
        font=button_font
    )
    style.map('Warning.TButton',
        background=[
            ('pressed', COLORS['warning']), 
            ('active', COLORS['warning']),
            ('disabled', COLORS['text_muted'])
        ]
    )

    # Success Button
    style.configure('Success.TButton', 
        background=COLORS['success'], 
        foreground='#ffffff', 
        borderwidth=0, 
        relief='flat',
        padding=(18, 16), 
        font=button_font
    )
    style.map('Success.TButton',
        background=[
            ('pressed', COLORS['success']), 
            ('active', COLORS['success']),
            ('disabled', COLORS['text_muted'])
        ]
    )

    # Modern Checkboxes and Radio buttons
    style.configure('TCheckbutton', 
        background=COLORS['card'], 
        foreground=COLORS['text'],
        font=base_font,
        focuscolor=COLORS['accent'])
    style.map('TCheckbutton',
        background=[('active', COLORS['card_hover'])],
        indicatorcolor=[('selected', COLORS['accent'])])
        
    style.configure('TRadiobutton', 
        background=COLORS['card'], 
        foreground=COLORS['text'],
        font=base_font,
        focuscolor=COLORS['accent'])
    style.map('TRadiobutton',
        background=[('active', COLORS['card_hover'])],
        indicatorcolor=[('selected', COLORS['accent'])])

    # Subtle modern scrollbars
    style.configure('TScrollbar',
        background=COLORS['border_light'],
        bordercolor=COLORS['border'],
        arrowcolor=COLORS['text_muted'],
        relief='flat',
        borderwidth=0,
        width=12)
    style.map('TScrollbar',
        background=[
            ('active', COLORS['border']),
            ('pressed', COLORS['border_light'])
        ],
        arrowcolor=[
            ('active', COLORS['accent']),
            ('pressed', COLORS['secondary'])
        ]
    )

    # Global TK widget styling
    root.option_add('*Font', base_font)
    root.option_add('*Button.relief', 'solid')
    root.option_add('*Button.borderWidth', '2')
    root.option_add('*Button.highlightThickness', '0')
    root.option_add('*Button.activeBackground', COLORS['card_hover'])
    root.option_add('*Button.activeForeground', COLORS['text'])
    root.option_add('*Button.background', COLORS['card'])
    root.option_add('*Button.foreground', COLORS['text'])
    root.option_add('*Button.font', ('Segoe UI', 10, 'bold'))

    # Modern window styling
    root.configure(background=COLORS['background'])
    
    # Try to enable modern window decorations (Windows 11 style)
    try:
        # Set modern window appearance
        root.tk.call('tk', 'scaling', 1.0)
    except Exception:
        pass
    
    return style


def style_scrolled_text(w):
    """ScrolledText/Text koyu kart görünümü için yardımcı."""
    try:
        w.configure(
            bg=COLORS['card'],
            fg=COLORS['text'],
            insertbackground=COLORS['accent'],
            relief='flat',
            bd=0,
            highlightthickness=1,
            highlightbackground=COLORS['border'],
            selectbackground=COLORS['hover'],
            selectforeground=COLORS['text'],
            padx=8, pady=6
        )
    except Exception:
        pass


def style_listbox(w):
    """Listbox koyu görünüm."""
    try:
        w.configure(
            bg=COLORS['card'], fg=COLORS['text'],
            selectbackground=COLORS['secondary'], selectforeground='#ffffff',
            relief='flat', bd=0, highlightthickness=1, highlightbackground=COLORS['border'],
            activestyle='none'
        )
    except Exception:
        pass


# ---------------------------------------------------------------------------
# ÖNEMLİ: Anahtar adlarını STANDARTLAŞTIRDIK
# 'e-Fatura Gelen' | 'e-Fatura Giden' | 'e-Arşiv Giden'
# --------------------------------------------------------------------------- #

# ReportModule uçları
REPORT_MODULES = {
    "Tümü": None,
    "e-Arşiv": {
        "templates_url": "https://api.nes.com.tr/earchive/v1/outgoing/reportmodule/templates",
        "create_url":    "https://api.nes.com.tr/earchive/v1/outgoing/reportmodule/reports",
        "download_url":  "https://api.nes.com.tr/earchive/v1/outgoing/reportmodule/reports/{id}/download"
    },
    "e-Fatura Gelen": {
        "templates_url": "https://api.nes.com.tr/einvoice/v1/incoming/reportmodule/templates",
        "create_url":    "https://api.nes.com.tr/einvoice/v1/incoming/reportmodule/reports",
        "download_url":  "https://api.nes.com.tr/einvoice/v1/incoming/reportmodule/reports/{id}/download"
    },
    "e-Fatura Giden": {
        "templates_url": "https://api.nes.com.tr/einvoice/v1/outgoing/reportmodule/templates",
        "create_url":    "https://api.nes.com.tr/einvoice/v1/outgoing/reportmodule/reports",
        "download_url":  "https://api.nes.com.tr/einvoice/v1/outgoing/reportmodule/reports/{id}/download"
    }
}

# Fatura temel uçları
BASE_URLS = {
    "Hepsi": None,
    "e-Fatura Gelen": "https://api.nes.com.tr/einvoice/v1/incoming/invoices",
    "e-Fatura Giden": "https://api.nes.com.tr/einvoice/v1/outgoing/invoices",
    "e-Arşiv Giden":  "https://api.nes.com.tr/earchive/v1/invoices"
}


# ===========================================================================
# Ayar / API yardımcıları
# ===========================================================================

def load_settings():
    s = APP_SETTINGS.copy()
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                s.update(json.load(f))
        except Exception:
            pass
    return s


def save_settings(settings: dict):
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)
    except Exception:
        pass


def load_apis():
    apis = {}
    # Registry
    if HAS_WINREG:
        try:
            k = winreg.OpenKey(winreg.HKEY_CURRENT_USER, REG_PATH)
            i = 0
            while True:
                name, val, _ = winreg.EnumValue(k, i)
                apis[name] = val
                i += 1
        except Exception:
            pass
    # JSON
    if os.path.exists(APIS_FILE):
        try:
            with open(APIS_FILE, 'r', encoding='utf-8') as f:
                apis.update(json.load(f))
        except Exception:
            pass
    return apis


def save_apis(apis: dict):
    # JSON
    try:
        with open(APIS_FILE, 'w', encoding='utf-8') as f:
            json.dump(apis, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

    # Registry
    if HAS_WINREG:
        try:
            # Eski anahtarı temizle
            try:
                winreg.DeleteKey(winreg.HKEY_CURRENT_USER, REG_PATH)
            except Exception:
                pass
            key = winreg.CreateKey(winreg.HKEY_CURRENT_USER, REG_PATH)
            for n, t in apis.items():
                winreg.SetValueEx(key, n, 0, winreg.REG_SZ, t)
            winreg.CloseKey(key)
        except Exception:
            pass


def apply_api(token: str, headers: dict):
    headers.clear()
    headers.update(HEADERS_TEMPLATE)
    headers['Authorization'] = f"Bearer {token}" if token else ""


# ===========================================================================
# Ortak yardımcılar – fatura/adlandırma/rapor çözümleme
# ===========================================================================

def sanitize_filename(header: str, fallback: str) -> str:
    """
    Content-Disposition header’ından dosya adını çıkarır; yoksa fallback kullanır.
    """
    m = re.search(r"filename\*\s*=\s*UTF-8''([^;]+)", header or '')
    if m:
        from urllib.parse import unquote
        return unquote(m.group(1))
    m = re.search(r'filename\s*=\s*"?([^";]+)"?', header or '')
    return m.group(1) if m else fallback


_TR_MAP = str.maketrans("ÇĞİÖŞÜ", "CGIOSU")
def _norm(s): return str(s).upper().translate(_TR_MAP).replace("\n", " ").strip()

COLUMN_MAP = {
    "tax": ["VERGI TUTARI", "KDV TUTARI", "KDV TOPLAMI", "HESAPLANAN KDV"],
    "payable": ["ODENECEK TUTAR", "MAL HIZMET TOPLAMI", "GENEL TOPLAM", "TOPLAM TUTAR"],
    "status": ["FATURA CEVABI", "KAYIT DURUMU", "BELGE DURUMU"]
}


def _find_idx(row, keys):
    for i, c in enumerate(row):
        if any(k in _norm(c) for k in keys):
            return i
    return None


def _to_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    m = re.search(r"([,\.])(\d{1,2})$", s)
    frac, int_pt = (m.group(2), s[:m.start(1)]) if m else ("", s)
    int_pt = re.sub(r"[ .,\u00A0]", "", int_pt)
    try:
        return float(int_pt + ("." + frac if frac else ""))
    except ValueError:
        return 0.0


def parse_report(path, mod):
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        header = hdr_row = None
        for r in range(1, 200):
            row = [c.value for c in ws[r]]
            if _find_idx(row, COLUMN_MAP["tax"]) is not None and \
               _find_idx(row, COLUMN_MAP["payable"]) is not None:
                header, hdr_row = row, r
                break
        if header is None:
            continue

        i_tax  = _find_idx(header, COLUMN_MAP["tax"])
        i_pay  = _find_idx(header, COLUMN_MAP["payable"])
        # Yalnız e-Fatura Giden için reddedilenleri düş
        is_giden = "fatura" in (mod or "").lower() and "giden" in (mod or "").lower()
        i_stat = _find_idx(header, COLUMN_MAP["status"]) if is_giden else None

        tax_sum = pay_sum = rej_sum = 0.0
        for row in ws.iter_rows(min_row=hdr_row+1, values_only=True):
            if not any(row):
                break
            tax = _to_float(row[i_tax]) if i_tax is not None else 0
            pay = _to_float(row[i_pay]) if i_pay is not None else 0
            if i_stat is not None and ("REDD" in _norm(row[i_stat] or "") or "REJECT" in _norm(row[i_stat] or "")):
                rej_sum += pay
                continue
            tax_sum += tax
            pay_sum += pay
        return tax_sum, pay_sum, rej_sum
    return 0.0, 0.0, 0.0


def annotate_excel(path, mod, year_month):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    meta = [
        ["Rapor Modülü",     mod],
        ["Dönem (YYYY-MM)",  year_month],
        ["Oluşturma Zamanı", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    ws.insert_rows(1, 3)
    for i, row in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=row[0])
        ws.cell(row=i, column=2, value=row[1])
    wb.save(path)


def _sanitize_name_part(s: str) -> str:
    s = str(s or "").strip()
    s = re.sub(r"[^\w\s\-\.]", "_", s, flags=re.UNICODE)
    s = re.sub(r"\s+", "_", s)
    return s[:80] if s else ""


def parse_invoice_list_text(text: str) -> dict:
    """
    Satır formatları:
      FATURANO
      FATURANO;ISIM
      FATURANO,ISIM
      FATURANO<TAB>ISIM
    Boş / yorum satırları (#) yok sayılır.
    Geriye {docNo: name_or_empty} dict döner.
    """
    mapping = {}
    if not text:
        return mapping
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        parts = re.split(r"[;,|\t]", line, maxsplit=1)
        doc = (parts[0] or "").strip()
        name = (parts[1] if len(parts) > 1 else "").strip()
        if doc:
            mapping[doc] = name
    return mapping


def parse_invoice_list_file(path: str) -> dict:
    """
    TXT/CSV => ilk iki sütun: [FATURANO, ISIM?]
    XLSX    => ilk iki anlamlı sütun: fatura no & isim; başlık aramaz.
    """
    ext = os.path.splitext(path)[1].lower()
    result = {}
    try:
        if ext in [".txt", ".csv"]:
            with open(path, "r", encoding="utf-8") as f:
                return parse_invoice_list_text(f.read())
        elif ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
            df = pd.read_excel(path, header=None, engine="openpyxl")
            if df is None or df.empty:
                return {}
            for _, row in df.iterrows():
                doc = str(row.iloc[0]).strip() if len(row) > 0 else ""
                name = str(row.iloc[1]).strip() if len(row) > 1 else ""
                if doc and doc.lower() not in ["nan", "none", ""]:
                    result[doc] = "" if (name or "").lower() in ["nan", "none"] else name
            return result
        else:
            return {}
    except Exception:
        return {}


# ===========================================================================
# API Manager (Ekle/Sil)
# ===========================================================================

class APIManager:
    def __init__(self, parent, apis, headers, on_change=None):
        self.parent  = parent
        self.apis    = apis
        self.headers = headers
        self.on_change = on_change

    def open_manager(self):
        mgr = tk.Toplevel(self.parent)
        mgr.title('API Yönetimi')
        mgr.geometry('420x360')
        mgr.resizable(False, False)
        mgr.grab_set()

        add = ttk.LabelFrame(mgr, text='Yeni API Ekle', padding=10)
        add.pack(fill='x', padx=10, pady=10)

        ttk.Label(add, text='API Adı:').grid(row=0, column=0, sticky='w', pady=5)
        name = tk.StringVar()
        ttk.Entry(add, textvariable=name).grid(row=0, column=1, sticky='ew', padx=5, pady=5)

        ttk.Label(add, text='Token:').grid(row=1, column=0, sticky='w', pady=5)
        tok = tk.StringVar()
        ttk.Entry(add, textvariable=tok, show='•').grid(row=1, column=1, sticky='ew', padx=5, pady=5)

        ttk.Button(add, text='✅  Ekle', style='Accent.TButton',
                   command=lambda: self._add_api(name, tok, mgr)).grid(row=2, column=1, sticky='e', pady=PADDING['medium'])

        # Liste
        lst_f = ttk.LabelFrame(mgr, text='Mevcut API\'ler', padding=10)
        lst_f.pack(fill='both', expand=True, padx=10, pady=10)

        api_list = tk.Listbox(lst_f)
        api_list.pack(side='left', fill='both', expand=True)
        sb = ttk.Scrollbar(lst_f, orient='vertical', command=api_list.yview)
        sb.pack(side='right', fill='y')
        api_list.config(yscrollcommand=sb.set)

        # Listbox stil ve kayıt
        try:
            style_listbox(api_list)
            top = mgr.winfo_toplevel()
            if hasattr(top, 'register_listbox_widget'):
                top.register_listbox_widget(api_list)
        except Exception:
            pass

        for api in self.apis:
            api_list.insert(tk.END, api)

        # Modern action buttons
        btn_f = ttk.Frame(mgr, style='Card.TFrame', padding=PADDING['medium'])
        btn_f.pack(fill='x')
        
        # Left side buttons
        left_buttons = ttk.Frame(btn_f, style='Card.TFrame')
        left_buttons.pack(side='left')
        
        ttk.Button(left_buttons, text='👁️  Görüntüle', style='Secondary.TButton',
                   command=lambda: self._view_api(
                       api_list.get(api_list.curselection()[0]) if api_list.curselection() else None
                   )).pack(side='left', padx=PADDING['xs'])
        
        ttk.Button(left_buttons, text='✏️  Düzenle', style='TButton',
                   command=lambda: self._edit_api(
                       api_list.get(api_list.curselection()[0]) if api_list.curselection() else None, mgr
                   )).pack(side='left', padx=PADDING['xs'])
        
        ttk.Button(left_buttons, text='🗑  Sil', style='Warning.TButton',
                   command=lambda: self._del_api(
                       api_list.get(api_list.curselection()[0]) if api_list.curselection() else None, mgr
                   )).pack(side='left', padx=PADDING['xs'])
        
        # Right side button
        ttk.Button(btn_f, text='✖  Kapat', style='TButton', command=mgr.destroy).pack(side='right', padx=PADDING['small'])

    def _add_api(self, name, tok, win):
        n, t = name.get().strip(), tok.get().strip()
        if not n or not t:
            messagebox.showerror('Hata', 'İsim ve token girin.')
            return
        self.apis[n] = t
        save_apis(self.apis)
        win.destroy()
        messagebox.showinfo('Başarılı', f"API '{n}' eklendi.")
        if self.on_change:
            self.on_change()

    def _view_api(self, name):
        """View API token in a secure dialog"""
        if not name:
            messagebox.showerror('Hata', 'Lütfen görüntülenecek API\'yi seçin.')
            return
        
        token = self.apis.get(name, '')
        if not token:
            messagebox.showerror('Hata', 'API token bulunamadı.')
            return
        
        # Create view dialog
        view_win = tk.Toplevel(self.parent)
        view_win.title(f'API Görüntüleme: {name}')
        view_win.geometry('500x200')
        view_win.resizable(False, False)
        view_win.grab_set()
        
        # Content frame
        content = ttk.LabelFrame(view_win, text=f'📋 {name} API Detayları', padding=PADDING['medium'])
        content.pack(fill='both', expand=True, padx=PADDING['medium'], pady=PADDING['medium'])
        
        # API Name
        ttk.Label(content, text='API Adı:', style='Card.TLabel', font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        ttk.Label(content, text=name, style='Card.TLabel').pack(anchor='w', pady=(0, PADDING['small']))
        
        # Token display with show/hide functionality
        ttk.Label(content, text='Token:', style='Card.TLabel', font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        
        token_frame = ttk.Frame(content, style='Card.TFrame')
        token_frame.pack(fill='x', pady=(0, PADDING['medium']))
        
        token_var = tk.StringVar(value='•' * 20)  # Hidden by default
        token_entry = ttk.Entry(token_frame, textvariable=token_var, state='readonly', font=('Consolas', 9))
        token_entry.pack(side='left', fill='x', expand=True, padx=(0, PADDING['small']))
        
        show_var = tk.BooleanVar()
        def toggle_token():
            if show_var.get():
                token_var.set(token)
                toggle_btn.config(text='🙈')
            else:
                token_var.set('•' * 20)
                toggle_btn.config(text='👁️')
        
        toggle_btn = ttk.Button(token_frame, text='👁️', width=4, command=toggle_token)
        toggle_btn.pack(side='right')
        
        # Close button
        ttk.Button(content, text='✖  Kapat', style='TButton', command=view_win.destroy).pack(pady=PADDING['small'])
    
    def _edit_api(self, name, parent_win):
        """Edit existing API"""
        if not name:
            messagebox.showerror('Hata', 'Lütfen düzenlenecek API\'yi seçin.')
            return
        
        current_token = self.apis.get(name, '')
        
        # Create edit dialog
        edit_win = tk.Toplevel(self.parent)
        edit_win.title(f'API Düzenleme: {name}')
        edit_win.geometry('450x180')
        edit_win.resizable(False, False)
        edit_win.grab_set()
        
        # Edit form
        edit_frame = ttk.LabelFrame(edit_win, text=f'✏️ {name} Düzenle', padding=PADDING['medium'])
        edit_frame.pack(fill='both', expand=True, padx=PADDING['medium'], pady=PADDING['medium'])
        
        # New token entry
        ttk.Label(edit_frame, text='Yeni Token:', style='Card.TLabel').pack(anchor='w')
        new_token_var = tk.StringVar(value=current_token)
        new_token_entry = ttk.Entry(edit_frame, textvariable=new_token_var, show='•', font=('Consolas', 9))
        new_token_entry.pack(fill='x', pady=(PADDING['xs'], PADDING['medium']))
        new_token_entry.focus()
        
        def save_changes():
            new_token = new_token_var.get().strip()
            if not new_token:
                messagebox.showerror('Hata', 'Token boş olamaz.')
                return
            
            self.apis[name] = new_token
            save_apis(self.apis)
            edit_win.destroy()
            messagebox.showinfo('Başarılı', f"API '{name}' güncellendi.")
            if self.on_change:
                self.on_change()
        
        # Buttons
        btn_frame = ttk.Frame(edit_frame, style='Card.TFrame')
        btn_frame.pack(fill='x')
        
        ttk.Button(btn_frame, text='💾  Kaydet', style='Accent.TButton', command=save_changes).pack(side='left')
        ttk.Button(btn_frame, text='✖  İptal', style='TButton', command=edit_win.destroy).pack(side='right')
        
        # Enter key binding
        edit_win.bind('<Return>', lambda e: save_changes())
    
    def _del_api(self, name, win):
        if not name:
            messagebox.showerror('Hata', 'Lütfen silinecek API\'yi seçin.')
            return
        if messagebox.askyesno('Onay', f"{name} silinsin mi?"):
            del self.apis[name]
            save_apis(self.apis)
            win.destroy()
            messagebox.showinfo('Başarılı', f"API '{name}' silindi.")
            if self.on_change:
                self.on_change()


# ===========================================================================
# ReportModule – Şablon listeleme
# ===========================================================================

def list_templates(mod, headers):
    if requests is None:
        raise RuntimeError("requests kütüphanesi yüklü değil. 'pip install requests'")

    if mod == 'Tümü':
        out = {}
        for m, cfg in REPORT_MODULES.items():
            if m == 'Tümü' or cfg is None:
                continue
            r = requests.get(cfg['templates_url'], headers=headers)
            r.raise_for_status()
            out[m] = r.json().get('data')
        return out

    cfg = REPORT_MODULES.get(mod)
    if cfg is None:
        return []
    r = requests.get(cfg['templates_url'], headers=headers)
    r.raise_for_status()
    return r.json().get('data')


# ===========================================================================
# Sekme: Fatura
# ===========================================================================

class InvoiceTab(ttk.Frame):
    def __init__(self, parent, apis, headers):
        super().__init__(parent, padding=10)
        self.apis = apis
        self.headers = headers
        self.settings = load_settings()

        self.save_folder = self.settings.get('invoice_folder') or ""
        self.stop_event = threading.Event()
        self.worker = None

        # Listeye göre indirme / adlandırma
        self.filter_enabled = tk.BooleanVar(value=False)
        self.use_name_in_filename = tk.BooleanVar(value=True)
        self.direct_number_mode = tk.BooleanVar(value=True)  # AY FİLTRESİ YOK → NUMARADAN İNDİR
        self.invoice_list_text = None  # scrolledtext
        self.invoice_list_map = {}     # {docNo: name}
        self.latest_output_folder = None

        for c in range(12):
            self.columnconfigure(c, weight=1)

        ttk.Label(self, text='API Hesabı:').grid(row=0, column=0, sticky='w')
        self.api_var = tk.StringVar()
        self.api_cb = ttk.Combobox(self, values=list(apis.keys()), textvariable=self.api_var, state='readonly')
        self.api_cb.grid(row=0, column=1, sticky='ew')
        if apis:
            last_api = self.settings.get('last_api', '')
            if last_api in apis:
                self.api_cb.set(last_api)
            else:
                self.api_cb.current(0)
            apply_api(apis[self.api_cb.get()], self.headers)
        else:
            self.api_cb.set('— API ekleyin —')

        # API Manager
        self.api_manager = APIManager(self, apis, headers, self._update_api_list)
        ttk.Button(self, text='API Ekle/Sil', style='Secondary.TButton', command=self.api_manager.open_manager).grid(row=0, column=2, sticky='ew', padx=PADDING['small'])

        # Klasör seçimi
        ttk.Button(self, text='Klasör Seç', style='TButton', command=self._select_folder).grid(row=0, column=3, sticky='ew', padx=PADDING['small'])
        self.folder_label = ttk.Label(self, text=f'Klasör: {self.save_folder or "Henüz seçilmedi"}')
        self.folder_label.grid(row=0, column=4, columnspan=8, sticky='w')

        # Tarih çerçevesi
        now = datetime.now()
        date_frame = ttk.Frame(self, style='Card.TFrame')
        date_frame.grid(row=1, column=0, columnspan=12, sticky='ew', pady=PADDING['small'])  # Add pady=PADDING['small']

        # Date frame labels
        ttk.Label(date_frame, text='Yıl:', style='Card.TLabel').grid(row=0, column=0, padx=PADDING['medium'])
        years = [str(now.year + i) for i in range(-10, 15)]
        self.year = ttk.Combobox(date_frame, values=years, width=6, state='readonly')
        self.year.set(str(now.year))
        self.year.grid(row=0, column=1, padx=PADDING['small'])

        ttk.Label(date_frame, text='Ay:', style='Card.TLabel').grid(row=0, column=2, padx=PADDING['small'])
        self.month = ttk.Combobox(date_frame, values=[f'{i:02d}' for i in range(1, 13)], width=4, state='readonly')
        self.month.set(f'{now.month:02d}')
        self.month.grid(row=0, column=3, padx=PADDING['small'])

        ttk.Label(date_frame, text='Tip:', style='Card.TLabel').grid(row=0, column=4, padx=PADDING['small'])
        self.kind = ttk.Combobox(date_frame, values=list(BASE_URLS.keys()), state='readonly', width=16)
        self.kind.set('Hepsi')
        self.kind.grid(row=0, column=5, padx=PADDING['small'])

        fmt_frame = ttk.Frame(date_frame, style='Card.TFrame')
        fmt_frame.grid(row=0, column=6, padx=PADDING['medium'], sticky='e')
        self.pdf_var = tk.BooleanVar(value=True)
        self.xml_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(fmt_frame, text='PDF', variable=self.pdf_var).pack(side='left', padx=6)
        ttk.Checkbutton(fmt_frame, text='XML', variable=self.xml_var).pack(side='left', padx=6)

        # Liste paneli
        list_frame = ttk.LabelFrame(self, text="Fatura No Listesi (opsiyonel)", padding=10)
        list_frame.grid(row=2, column=0, columnspan=12, sticky='ew', pady=PADDING['small'])
        list_frame.columnconfigure(1, weight=1)
        list_frame.columnconfigure(2, weight=1)

        ttk.Checkbutton(list_frame, text="Listeye göre indir", variable=self.filter_enabled).grid(row=0, column=0, sticky='w', pady=2)
        ttk.Checkbutton(list_frame, text="Kaydetme adında ismi kullan (İSİM_FATURANO)", variable=self.use_name_in_filename).grid(row=0, column=1, sticky='w', pady=2)
        ttk.Checkbutton(list_frame, text="Ay filtresi olmadan, numaradan indir (yalnız listedekiler)", variable=self.direct_number_mode).grid(row=0, column=2, sticky='w', pady=2)

        ttk.Label(list_frame, text="Satır formatı: FATURANO;İSİM  (virgül/sekme kabul)", style='Card.TLabel').grid(row=1, column=0, columnspan=3, sticky='w')

        self.invoice_list_text = scrolledtext.ScrolledText(list_frame, height=6)
        self.invoice_list_text.grid(row=2, column=0, columnspan=3, sticky='ew', pady=6)
        style_scrolled_text(self.invoice_list_text)
        try:
            top = self.winfo_toplevel()
            if hasattr(top, 'register_text_widget'):
                top.register_text_widget(self.invoice_list_text)
        except Exception:
            pass

        btns = ttk.Frame(list_frame)
        btns.grid(row=3, column=0, columnspan=3, sticky='e', pady=PADDING['small'])
        ttk.Button(btns, text='Dosyadan Yükle', style='TButton', command=self._load_list_file).pack(side='left', padx=PADDING['small'])
        ttk.Button(btns, text='Listeyi Temizle', style='Warning.TButton', command=lambda: self.invoice_list_text.delete('1.0', tk.END)).pack(side='left', padx=PADDING['small'])
        ttk.Button(btns, text='No ile Sorgula → Excel', style='Secondary.TButton', command=self._query_numbers_to_excel).pack(side='left', padx=PADDING['small'])

        # Modern Action Buttons
        btn_row = ttk.Frame(self, style='Card.TFrame', padding=PADDING['medium'])
        btn_row.grid(row=3, column=0, columnspan=12, sticky='ew', pady=PADDING['medium'])
        
        # Primary action button
        self.btn_start = ttk.Button(btn_row, text='🚀  İndir & ZIP', style='Accent.TButton', command=self._start)
        self.btn_start.pack(side='right', padx=PADDING['small'])
        
        # Stop button
        self.btn_stop = ttk.Button(btn_row, text='⏹  Durdur', style='Warning.TButton', command=self._stop, state='disabled')
        self.btn_stop.pack(side='right', padx=PADDING['small'])
        
        # Secondary action button
        self.btn_open = ttk.Button(btn_row, text='📁  Klasöre Git', style='TButton', command=self._open_latest)
        self.btn_open.pack(side='right', padx=PADDING['small'])

        # Modern Log Section
        log_frame = ttk.LabelFrame(self, text='📋 Log Kayıtları', style='TLabelframe', padding=PADDING['medium'])
        log_frame.grid(row=4, column=0, columnspan=12, sticky='nsew', pady=PADDING['medium'])
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log = scrolledtext.ScrolledText(log_frame, height=18, wrap=tk.WORD)
        self.log.grid(row=0, column=0, sticky='nsew')
        style_scrolled_text(self.log)
        try:
            top = self.winfo_toplevel()
            if hasattr(top, 'register_text_widget'):
                top.register_text_widget(self.log)
        except Exception:
            pass

        self.rowconfigure(4, weight=1)

        self.api_cb.bind('<<ComboboxSelected>>', self._on_api_change)

    # -------------------- Liste / UI yardımcıları -------------------- #

    def _load_list_file(self):
        path = filedialog.askopenfilename(
            title="Fatura No listesi seçin",
            filetypes=[("Tümü", "*.txt *.csv *.xlsx *.xlsm *.xltx *.xltm"),
                       ("Metin", "*.txt"), ("CSV", "*.csv"), ("Excel", "*.xlsx *.xlsm *.xltx *.xltm")]
        )
        if not path:
            return
        mapping = parse_invoice_list_file(path)
        if not mapping:
            messagebox.showwarning("Uyarı", "Dosyadan anlamlı bir veri okunamadı.")
            return
        cur = self.invoice_list_text.get('1.0', tk.END).strip()
        rows = []
        if cur:
            rows.append(cur)
        for k, v in mapping.items():
            rows.append(f"{k};{v}" if v else f"{k}")
        self.invoice_list_text.delete('1.0', tk.END)
        self.invoice_list_text.insert(tk.END, "\n".join(rows))

    def _update_api_list(self):
        self.api_cb['values'] = list(self.apis.keys()) or ['— API ekleyin —']
        if self.apis and self.api_cb.get() not in self.apis:
            self.api_cb.current(0)
            apply_api(self.apis[self.api_cb.get()], self.headers)

    def _on_api_change(self, _=None):
        api_name = self.api_cb.get()
        if api_name in self.apis:
            apply_api(self.apis[api_name], self.headers)
            self.settings['last_api'] = api_name
            save_settings(self.settings)

    def _select_folder(self):
        d = filedialog.askdirectory(title='Kayıt Klasörü Seç')
        if d:
            self.save_folder = d
            self.folder_label.config(text=f'Klasör: {d}')
            self.settings['invoice_folder'] = d
            save_settings(self.settings)

    def _guard_api(self) -> bool:
        if requests is None:
            messagebox.showerror("Eksik Bağımlılık", "requests yüklü değil.\nKomut: pip install requests")
            return False
        api_name = self.api_cb.get()
        if api_name not in self.apis or not (self.apis.get(api_name) or "").strip():
            messagebox.showwarning("API Gerekli",
                                   "Bu işlem için API token gerekli. Lütfen 'API Ekle/Sil' ile bir hesap ekleyin.")
            return False
        return True

    # -------------------- İndir & ZIP akışı -------------------- #

    def _start(self):
        if not self.save_folder:
            messagebox.showerror('Hata', 'Lütfen kayıt klasörü seçin.')
            return
        # Numara modu ay/yıl istemez
        if not (self.filter_enabled.get() and self.direct_number_mode.get()):
            try:
                y = int(self.year.get()); m = int(self.month.get())
            except Exception:
                messagebox.showerror('Hata', 'Geçerli yıl/ay seçin.')
                return
            if self.kind.get() != 'Hepsi' and self.kind.get() not in BASE_URLS:
                messagebox.showerror('Hata', 'Geçerli tip seçin.')
                return
        if not self._guard_api():
            return

        self.invoice_list_map = {}
        if self.filter_enabled.get():
            txt = self.invoice_list_text.get('1.0', tk.END)
            self.invoice_list_map = parse_invoice_list_text(txt) or {}
            if not self.invoice_list_map:
                messagebox.showwarning("Uyarı", "‘Listeye göre indir’ aktif ama listede veri yok.")
                return

        self.log_delete()
        self.stop_event.clear()
        self.btn_start.config(state='disabled')
        self.btn_stop.config(state='normal')

        if self.filter_enabled.get() and self.direct_number_mode.get():
            t = threading.Thread(target=self._run_numbers_only, daemon=True)
        else:
            y = int(self.year.get()); m = int(self.month.get())
            t = threading.Thread(target=self._run, args=(y, m), daemon=True)

        self.worker = t
        t.start()

    def _stop(self):
        self.stop_event.set()
        self._log("❌ İndirme durdurma talebi alındı. Mevcut adım bitince duracak...")
        self.btn_stop.config(state='disabled')

    def _open_latest(self):
        if self.latest_output_folder and os.path.exists(self.latest_output_folder):
            webbrowser.open(self.latest_output_folder)
        else:
            messagebox.showinfo("Bilgi", "Henüz bir çıktı klasörü oluşmadı.")

    def _get_invoices(self, url, start_iso, end_iso, page_size=50):
        invs, page, total = [], 1, None
        while not self.stop_event.is_set():
            params = {
                "archived": "false", "sort": "createdAt desc",
                "pageSize": page_size, "page": page,
                "startDate": start_iso, "endDate": end_iso
            }
            r = requests.get(url, headers=self.headers, params=params)
            if r.status_code != 200:
                self._log(f"Hata {r.status_code}: {r.text}")
                break
            data = r.json()
            if total is None:
                total = math.ceil(data.get('totalCount', 0) / page_size) or 1
            batch = data.get('data') or data.get('invoices') or data.get('items') or data.get('content') or data.get('results') or data.get('list') or []
            if not batch:
                break
            invs.extend(batch)
            self._log(f"Sayfa {page}/{total} → {len(batch)} fatura")
            if page >= total:
                break
            page += 1
        return invs

    def _download_docs(self, key, invs, folder, pdf_opt, xml_opt, filter_map=None, use_name_in_filename=False):
        found = set()
        saved = 0
        base = os.path.join(folder, key.replace(' ', '_'))
        os.makedirs(base, exist_ok=True)

        for inv in invs:
            if self.stop_event.is_set():
                break
            num, uid = inv.get('documentNumber') or inv.get('id'), inv.get('id')
            if not num or not uid:
                continue

            if filter_map is not None and num not in filter_map:
                continue
            if filter_map is not None:
                found.add(num)

            # Dosya adı
            name_part = ""
            if use_name_in_filename and filter_map is not None:
                name_part = _sanitize_name_part(filter_map.get(num, "")) or ""
            base_name = f"{name_part+'_' if name_part else ''}{num}"

            if pdf_opt:
                d = os.path.join(base, 'pdf'); os.makedirs(d, exist_ok=True)
                r = requests.get(f"{BASE_URLS[key]}/{uid}/pdf", headers=self.headers)
                if r.status_code == 200:
                    with open(os.path.join(d, f"{base_name}.pdf"), 'wb') as f:
                        f.write(r.content)
                    saved += 1
            if xml_opt:
                d = os.path.join(base, 'xml'); os.makedirs(d, exist_ok=True)
                r = requests.get(f"{BASE_URLS[key]}/{uid}/xml", headers=self.headers)
                if r.status_code == 200:
                    with open(os.path.join(d, f"{base_name}.xml"), 'wb') as f:
                        f.write(r.content)

        self._log(f"İndirme tamamlandı ({key}). Kayıt: {saved} dosya.")

        # Bulunamayanlar raporu
        if filter_map is not None:
            missing = sorted(set(filter_map.keys()) - found)
            if missing:
                miss_path = os.path.join(base, "bulunamayanlar.txt")
                try:
                    with open(miss_path, "w", encoding="utf-8") as f:
                        for doc in missing:
                            nm = filter_map.get(doc, "")
                            f.write(f"{doc};{nm}\n")
                    self._log(f"⚠️ Bulunamayanlar kaydedildi: {miss_path}")
                except OSError as e:
                    self._log(f"⚠️ Bulunamayanlar yazılamadı: {e}")

    def _zip_folder(self, folder):
        zf_path = f"{folder}.zip"
        with zipfile.ZipFile(zf_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, _, files in os.walk(folder):
                for f in files:
                    full = os.path.join(root, f)
                    zf.write(full, os.path.relpath(full, folder))
        self._log(f"ZIP oluşturuldu: {zf_path}")

    def _run(self, y, m):
        api_name = self.api_cb.get()
        apply_api(self.apis.get(api_name, ''), self.headers)

        today = datetime.now().strftime('%d.%m.%Y')
        sd = datetime(y, m, 1).strftime('%Y-%m-%dT00:00:00+03:00')
        ed = datetime(y, m, calendar.monthrange(y, m)[1], 23, 59, 59).strftime('%Y-%m-%dT%H:%M:%S+03:00')

        folder_name = f"{api_name} {int(m)}. ay {today} Faturaları"
        base = os.path.join(self.save_folder, folder_name)
        os.makedirs(base, exist_ok=True)
        self.latest_output_folder = base

        self._log(f"▶ API Hesabı: {api_name}")
        self._log(f"▶ Fatura Başlatıldı: {y}-{m:02d} ({today})")
        self._log(f" • Çıktı klasörü: {base}")

        # Akıllı tarama
        if self.filter_enabled.get():
            prefixes = {str(doc)[:3].upper() for doc in self.invoice_list_map.keys() if doc}
            auto = []
            if 'EAR' in prefixes: auto.append('e-Arşiv Giden')
            if 'EFR' in prefixes: auto.append('e-Fatura Giden')
            kinds = auto or ([self.kind.get()] if self.kind.get() != 'Hepsi' else [k for k in BASE_URLS if BASE_URLS[k]])
        else:
            kinds = ([self.kind.get()] if self.kind.get() != 'Hepsi' else [k for k in BASE_URLS if BASE_URLS[k]])

        for k in kinds:
            if self.stop_event.is_set():
                break
            self._log(f"\n→ {k} çekiliyor...")
            invs = self._get_invoices(BASE_URLS[k], sd, ed)

            active_filter = self.invoice_list_map if self.filter_enabled.get() else None
            self._download_docs(
                k, invs, base, self.pdf_var.get(), self.xml_var.get(),
                filter_map=active_filter,
                use_name_in_filename=self.use_name_in_filename.get()
            )

        if not self.stop_event.is_set() and self.settings.get('auto_zip', True):
            self._zip_folder(base)

        self._log('▶ Fatura tamamlandı!' if not self.stop_event.is_set() else '▶ İşlem kullanıcı tarafından durduruldu.')
        self.btn_start.config(state='normal')
        self.btn_stop.config(state='disabled')

    # -------------------- NUMARADAN İNDİRME -------------------- #

    def _download_pdf_xml(self, key, inv, base, base_name, pdf_opt, xml_opt):
        uid = inv.get('id')
        if not uid:
            return 0
        saved = 0
        mod_dir = os.path.join(base, key.replace(' ', '_'))
        if pdf_opt:
            d = os.path.join(mod_dir, 'pdf'); os.makedirs(d, exist_ok=True)
            r = requests.get(f"{BASE_URLS[key]}/{uid}/pdf", headers=self.headers)
            if r.status_code == 200:
                with open(os.path.join(d, f"{base_name}.pdf"), 'wb') as f:
                    f.write(r.content)
                saved += 1
        if xml_opt:
            d = os.path.join(mod_dir, 'xml'); os.makedirs(d, exist_ok=True)
            r = requests.get(f"{BASE_URLS[key]}/{uid}/xml", headers=self.headers)
            if r.status_code == 200:
                with open(os.path.join(d, f"{base_name}.xml"), 'wb') as f:
                    f.write(r.content)
                saved += 1
        return saved

    def _run_numbers_only(self):
        api_name = self.api_cb.get()
        apply_api(self.apis.get(api_name, ''), self.headers)

        today = datetime.now().strftime('%d.%m.%Y')
        folder_name = f"{api_name} NumaraModu {today} Faturalar"
        base = os.path.join(self.save_folder, folder_name)
        os.makedirs(base, exist_ok=True)
        self.latest_output_folder = base

        self._log(f"▶ API Hesabı: {api_name}")
        self._log(f"▶ Numara Modu: Ay filtresi YOK (tarih bağımsız).")
        self._log(f" • Çıktı klasörü: {base}")

        mapping = self.invoice_list_map or {}
        missing = []
        total_saved = 0

        for doc, label in mapping.items():
            if self.stop_event.is_set():
                break
            doc = str(doc).strip()
            if not doc:
                continue

            # Önekten tahmin:
            pref = doc[:3].upper()
            if pref == "EAR":
                mods = ["e-Arşiv Giden", "e-Fatura Giden", "e-Fatura Gelen"]
            elif pref == "EFR":
                mods = ["e-Fatura Giden", "e-Fatura Gelen", "e-Arşiv Giden"]
            else:
                mods = ["e-Fatura Giden", "e-Fatura Gelen", "e-Arşiv Giden"]

            inv = found_mod = None
            for mod in mods:
                inv = self._search_invoice_by_number(mod, doc)
                if inv:
                    found_mod = mod
                    break

            if not inv:
                self._log(f"⚠️ Bulunamadı: {doc}")
                missing.append(f"{doc};{label or ''}")
                continue

            name_part = _sanitize_name_part(label) if label else ""
            base_name = f"{name_part+'_' if name_part else ''}{doc}"
            saved = self._download_pdf_xml(found_mod, inv, base, base_name, self.pdf_var.get(), self.xml_var.get())
            total_saved += saved
            self._log(f"✓ {doc} → {found_mod} ({saved} dosya)")

        # Bulunamayanlar
        if missing:
            miss_path = os.path.join(base, "bulunamayanlar.txt")
            try:
                with open(miss_path, "w", encoding="utf-8") as f:
                    f.write("\n".join(missing))
                self._log(f"⚠️ Bulunamayanlar kaydedildi: {miss_path}")
            except OSError as e:
                self._log(f"⚠️ Bulunamayanlar yazılamadı: {e}")

        if not self.stop_event.is_set() and self.settings.get('auto_zip', True):
            self._zip_folder(base)

        self._log(f"▶ Numara Modu tamamlandı. Toplam kayıt: {total_saved} dosya.")
        self.btn_start.config(state='normal')
        self.btn_stop.config(state='disabled')

    # -------------------- Numara Bazlı Sorgu → Excel -------------------- #

    def _deep_get(self, d, path):
        cur = d
        for p in path:
            if not isinstance(cur, dict):
                return None
            cur = cur.get(p)
        return cur

    def _any_of(self, d, paths):
        for p in paths:
            v = self._deep_get(d, p) if isinstance(p, (list, tuple)) else (d.get(p) if isinstance(d, dict) else None)
            if v not in (None, "", [], {}):
                return v
        return None

    def _fmt_date(self, v):
        if not v:
            return ""
        try:
            if isinstance(v, (int, float)):
                return str(v)
            s = str(v)
            if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                return s[:10]
            dt = datetime.fromisoformat(s.replace('Z','').replace('T',' '))
            return dt.strftime('%Y-%m-%d')
        except Exception:
            return str(v)

    def _search_invoice_by_number(self, mod: str, doc: str):
        """
        Belge numarasına göre tekil kaydı bulmaya çalışır.
        Hızlı parametre varyasyonları → by-number uçları → geniş tarama fallback.
        """
        if requests is None:
            return None

        base = BASE_URLS.get(mod)
        if not base:
            return None

        doc_s = str(doc).strip()

        # --- 1) Hızlı parametre varyasyonları
        variants = [
            {"documentNumber": doc_s},
            {"documentNumbers": doc_s},
            {"invoiceNumber": doc_s},
            {"search": doc_s},
            {"query": doc_s},
        ]
        for params in variants:
            try:
                rr = requests.get(base, headers=self.headers, params={**params, "pageSize": 5, "page": 1})
                if rr.status_code != 200:
                    continue
                data = rr.json() or {}

                # Tek obje
                if isinstance(data, dict):
                    num0 = (
                        data.get("documentNumber")
                        or data.get("number")
                        or self._deep_get(data, ["document", "number"])
                    )
                    if str(num0).strip() == doc_s:
                        return data

                # Liste
                items = (
                    data.get("data")
                    or data.get("invoices")
                    or data.get("items")
                    or data.get("content")
                    or data.get("results")
                    or data.get("list")
                    or []
                )
                for it in items:
                    num = (
                        it.get("documentNumber")
                        or it.get("number")
                        or self._deep_get(it, ["document", "number"])
                    )
                    if str(num).strip() == doc_s:
                        return it
            except Exception:
                continue

        # --- 2) Muhtemel by-number uçları
        for suffix in (f"/documentNumber/{doc_s}", f"/by-document-number/{doc_s}"):
            try:
                rr = requests.get(base + suffix, headers=self.headers)
                if rr.status_code == 200:
                    it = rr.json() or {}
                    num = (
                        it.get("documentNumber")
                        or it.get("number")
                        or self._deep_get(it, ["document", "number"])
                    )
                    if str(num).strip() == doc_s:
                        return it
            except Exception:
                pass

        # --- 3) Geniş tarama fallback
        self._log(f"ℹ️ Hızlı aramada bulunamadı: {doc_s}. → {mod} için geniş tarama başlıyor (36 ay).")
        return self._search_invoice_by_number_wide(mod, doc_s, months=36, step_days=90)

    def _search_invoice_by_number_wide(self, mod: str, doc: str, months: int = 36, step_days: int = 90):
        if requests is None:
            return None

        base = BASE_URLS.get(mod)
        if not base:
            return None

        end = datetime.now()
        start_limit = end - timedelta(days=months * 30)

        window_idx = 0
        while end > start_limit and not self.stop_event.is_set():
            start = max(start_limit, end - timedelta(days=step_days - 1))

            start_iso = start.strftime('%Y-%m-%dT00:00:00+03:00')
            end_iso   = end.strftime('%Y-%m-%dT23:59:59+03:00')

            window_idx += 1
            self._log(f"  🔎 [{mod}] pencere #{window_idx}: {start.strftime('%Y-%m-%d')} → {end.strftime('%Y-%m-%d')}")

            try:
                invs = self._get_invoices(base, start_iso, end_iso)
                for it in invs:
                    num = (
                        it.get("documentNumber")
                        or it.get("number")
                        or self._deep_get(it, ["document", "number"])
                    )
                    if str(num).strip() == str(doc).strip():
                        self._log(f"  ✅ {doc} bulundu ({mod}) → pencere #{window_idx}")
                        return it
            except Exception as e:
                self._log(f"  ⚠️ Geniş taramada hata: {e}")

            end = start - timedelta(days=1)

        self._log(f"  ❌ {doc} geniş taramada da bulunamadı ({mod}).")
        return None

    def _query_numbers_to_excel(self):
        if not self._guard_api():
            return
        txt = self.invoice_list_text.get('1.0', tk.END)
        mapping = parse_invoice_list_text(txt)  # {docNo: optional_name}
        if not mapping:
            messagebox.showwarning("Uyarı", "Fatura no listesi boş.")
            return

        rows = []
        for doc, _label in mapping.items():
            doc = str(doc).strip()
            if not doc:
                continue
            # Önekten modül tahmini
            pref = doc[:3].upper()
            if pref == "EAR":
                mods = ["e-Arşiv Giden", "e-Fatura Giden", "e-Fatura Gelen"]
            elif pref == "EFR":
                mods = ["e-Fatura Giden", "e-Fatura Gelen", "e-Arşiv Giden"]
            else:
                mods = ["e-Fatura Giden", "e-Fatura Gelen", "e-Arşiv Giden"]

            inv = found_mod = None
            for mod in mods:
                inv = self._search_invoice_by_number(mod, doc)
                if inv:
                    found_mod = mod
                    break

            if not inv:
                rows.append({
                    "SATIŞ TARİHİ": "",
                    "ADI SOYADI": "",
                    "SATIŞ BEDELİ": "",
                    "KDV TUTARI": "",
                    "SATIŞ BELGESİNİN NUMARASI": doc,
                    "KİMLİK NO/vergi no": "",
                    "Fatura Türü": "BULUNAMADI",
                })
                continue

            # Tarih
            date = self._any_of(inv, [
                ["document", "issueDate"], ["issueDate"], ["documentDate"], ["createdAt"]
            ])

            # Karşı taraf adı / kimlik
            if found_mod in ("e-Arşiv Giden", "e-Fatura Giden"):
                name = self._any_of(inv, [
                    ["receiver", "title"], ["receiver", "name"], ["receiverTitle"], ["buyer", "title"], ["buyer", "name"]
                ])
                ident = self._any_of(inv, [
                    ["receiver", "identifier"], ["receiver", "taxOrId"], ["receiver", "vknTckn"],
                    ["buyer", "identifier"], ["buyer", "taxId"]
                ])
            else:  # e-Fatura Gelen
                name = self._any_of(inv, [
                    ["supplier", "title"], ["supplier", "name"], ["sender", "title"], ["sender", "name"], ["supplierTitle"]
                ])
                ident = self._any_of(inv, [
                    ["supplier", "identifier"], ["supplier", "taxOrId"], ["supplier", "vknTckn"],
                    ["sender", "identifier"], ["sender", "taxId"]
                ])

            # Tutarlar
            total = self._any_of(inv, [
                ["totals", "payable"], ["summary", "grandTotal"], ["summary", "payableAmount"],
                ["payableAmount"], ["grandTotal"], ["totalAmount"], ["amount"]
            ])
            tax = self._any_of(inv, [
                ["totals", "tax"], ["taxTotal", "taxAmount"], ["summary", "vatAmount"], ["vatTotal"]
            ])

            def _as_float(x):
                try:
                    return _to_float(x)
                except Exception:
                    return 0.0

            rows.append({
                "SATIŞ TARİHİ": self._fmt_date(date),
                "ADI SOYADI": name or "",
                "SATIŞ BEDELİ": _as_float(total),
                "KDV TUTARI": _as_float(tax),
                "SATIŞ BELGESİNİN NUMARASI": (inv.get("documentNumber") or self._deep_get(inv, ["document","number"]) or doc),
                "KİMLİK NO/vergi no": ident or "",
                "Fatura Türü": found_mod,
            })

        # Excel çıktısı
        out_dir = self.save_folder or os.getcwd()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = os.path.join(out_dir, f"sorgu_{ts}.xlsx")
        try:
            pd.DataFrame(rows, columns=[
                "SATIŞ TARİHİ","ADI SOYADI","SATIŞ BEDELİ","KDV TUTARI",
                "SATIŞ BELGESİNİN NUMARASI","KİMLİK NO/vergi no","Fatura Türü"
            ]).to_excel(out_path, index=False)
            self._log(f"📊 Sorgu raporu oluşturuldu: {out_path}")
            webbrowser.open(os.path.dirname(out_path))
        except Exception as e:
            messagebox.showerror("Hata", f"Excel yazılamadı:\n{e}")

    # -------------------- Log yardımcıları -------------------- #

    def _log(self, msg):
        self.log.insert(tk.END, msg + '\n')
        self.log.see(tk.END)

    def log_delete(self):
        self.log.delete('1.0', tk.END)


# ===========================================================================
# Sekme: Rapor (ReportModule + KDV özeti)
# ===========================================================================

class ReportTab(ttk.Frame):
    def __init__(self, parent, apis, headers):
        super().__init__(parent, padding=PADDING['large'])
        self.apis = apis
        self.headers = headers
        self.settings = load_settings()
        self.sums = {}
        self.stop_event = threading.Event()
        self.worker = None
        self.output_dir = None

        # Modern control panel with proper spacing
        control_frame = ttk.LabelFrame(self, text='⚙️ Rapor Ayarları', style='TLabelframe', padding=PADDING['medium'])
        control_frame.pack(fill='x', pady=(0, PADDING['medium']))
        control_frame.columnconfigure(1, weight=1)
        control_frame.columnconfigure(3, weight=1)

        # Row 1: API and Module
        ttk.Label(control_frame, text='API Hesabı:', style='Card.TLabel', font=('Segoe UI', 11)).grid(row=0, column=0, sticky='w', padx=(0, PADDING['small']), pady=PADDING['xs'])
        self.api_var = tk.StringVar()
        self.api_cb = ttk.Combobox(control_frame, values=list(apis.keys()), textvariable=self.api_var, state='readonly', width=20)
        self.api_cb.grid(row=0, column=1, sticky='ew', padx=(0, PADDING['medium']), pady=PADDING['xs'])
        
        if apis:
            last_api = self.settings.get('last_api', '')
            if last_api in apis:
                self.api_cb.set(last_api)
            else:
                self.api_cb.current(0)
            apply_api(apis[self.api_cb.get()], self.headers)
        else:
            self.api_cb.set('— API ekleyin —')

        self.api_manager = APIManager(self, apis, headers, self._update_api_list)
        ttk.Button(control_frame, text='🔧 API Yönetimi', style='Secondary.TButton', 
                  command=self.api_manager.open_manager).grid(row=0, column=2, sticky='w', padx=PADDING['small'], pady=PADDING['xs'])

        ttk.Label(control_frame, text='Modül:', style='Card.TLabel', font=('Segoe UI', 11)).grid(row=0, column=3, sticky='w', padx=(PADDING['large'], PADDING['small']), pady=PADDING['xs'])
        self.mod_cb = ttk.Combobox(control_frame, values=list(REPORT_MODULES.keys()), state='readonly', width=18)
        self.mod_cb.set('Tümü')
        self.mod_cb.grid(row=0, column=4, sticky='ew', pady=PADDING['xs'])

        # Row 2: Date selection
        date_frame = ttk.Frame(control_frame, style='Card.TFrame')
        date_frame.grid(row=1, column=0, columnspan=5, sticky='ew', pady=(PADDING['small'], 0))
        
        now = datetime.now()
        ttk.Label(date_frame, text='Dönem:', style='Card.TLabel', font=('Segoe UI', 11, 'bold')).pack(side='left', padx=(0, PADDING['medium']))
        
        ttk.Label(date_frame, text='Yıl:', style='Card.TLabel').pack(side='left', padx=(0, PADDING['xs']))
        years = [str(now.year + i) for i in range(-10, 15)]
        self.year = ttk.Combobox(date_frame, values=years, width=8, state='readonly')
        self.year.set(str(now.year))
        self.year.pack(side='left', padx=(0, PADDING['medium']))

        ttk.Label(date_frame, text='Ay:', style='Card.TLabel').pack(side='left', padx=(0, PADDING['xs']))
        self.month = ttk.Combobox(date_frame, values=[f'{i:02d}' for i in range(1, 13)], width=6, state='readonly')
        self.month.set(f'{now.month:02d}')
        self.month.pack(side='left', padx=(0, PADDING['large']))

        self.btn_start = ttk.Button(date_frame, text='🚀 Rapor Oluştur', style='Accent.TButton', command=self._start)
        self.btn_start.pack(side='right', padx=(PADDING['medium'], 0))

        # Modern Log Section
        log_frame = ttk.LabelFrame(self, text='📋 Log Kayıtları', style='TLabelframe', padding=PADDING['medium'])
        log_frame.pack(fill='both', expand=True, pady=PADDING['medium'])
        
        self.log = scrolledtext.ScrolledText(log_frame, height=20, wrap=tk.WORD, font=('Consolas', 10))
        self.log.pack(fill='both', expand=True)
        style_scrolled_text(self.log)
        try:
            top = self.winfo_toplevel()
            if hasattr(top, 'register_text_widget'):
                top.register_text_widget(self.log)
        except Exception:
            pass

        # Configure main layout
        self.pack_propagate(True)
        self.api_cb.bind('<<ComboboxSelected>>', self._on_api_change)
        self.mod_cb.bind('<<ComboboxSelected>>', self._load_templates)

        # Modern Template Section
        template_frame = ttk.LabelFrame(self, text='📄 Şablon Ayarları', style='TLabelframe', padding=PADDING['medium'])
        template_frame.pack(fill='x', pady=PADDING['small'])
        template_frame.columnconfigure(1, weight=1)
        
        # Template fields
        self.tpl = tk.StringVar()
        self.tpl_cb = ttk.Combobox(template_frame, textvariable=self.tpl, state='readonly', font=('Segoe UI', 10))
        
        ttk.Label(template_frame, text='Şablon:', style='Card.TLabel', font=('Segoe UI', 11)).grid(row=0, column=0, sticky='w', padx=(0, PADDING['small']), pady=PADDING['xs'])
        self.tpl_cb.grid(row=0, column=1, sticky='ew', padx=(0, PADDING['medium']), pady=PADDING['xs'])

        ttk.Label(template_frame, text='Başlık:', style='Card.TLabel', font=('Segoe UI', 11)).grid(row=1, column=0, sticky='w', padx=(0, PADDING['small']), pady=PADDING['xs'])
        self.title = ttk.Entry(template_frame, font=('Segoe UI', 11))
        self.title.insert(0, 'Rapor')
        self.title.grid(row=1, column=1, sticky='ew', padx=(0, PADDING['medium']), pady=PADDING['xs'])

        # Action button
        action_frame = ttk.Frame(template_frame, style='Card.TFrame')
        action_frame.grid(row=2, column=0, columnspan=2, sticky='ew', pady=(PADDING['small'], 0))
        
        self.btn_open = ttk.Button(action_frame, text='📁  Klasöre Git', style='TButton', command=self._open_output)
        self.btn_open.pack(side='right')

        # İlk yükle
        try:
            self._load_templates()
        except Exception as e:
            messagebox.showwarning('Şablon Yüklenemedi', f'Şablonlar yüklenirken hata oluştu:\n{e}')

    def _update_api_list(self):
        self.api_cb['values'] = list(self.apis.keys()) or ['— API ekleyin —']
        if self.apis and self.api_cb.get() not in self.apis:
            self.api_cb.current(0)
            apply_api(self.apis[self.api_cb.get()], self.headers)

    def _on_api_change(self, _=None):
        api_name = self.api_cb.get()
        if api_name in self.apis:
            apply_api(self.apis[api_name], self.headers)
            self.settings['last_api'] = api_name
            save_settings(self.settings)
            # Şablonları yenile
            self._load_templates()

    def _guard_api(self) -> bool:
        if requests is None:
            messagebox.showerror("Eksik Bağımlılık", "requests yüklü değil.\nKomut: pip install requests")
            return False
        api_name = self.api_cb.get()
        if api_name not in self.apis or not (self.apis.get(api_name) or "").strip():
            messagebox.showwarning("API Gerekli",
                                   "Bu işlem için API token gerekli. Lütfen 'API Ekle/Sil' ile bir hesap ekleyin.")
            return False
        return True

    def _load_templates(self):
        if not self._guard_api():
            self.tpl_cb['values'] = []
            return
        items = list_templates(self.mod_cb.get(), self.headers)
        vals = []
        if isinstance(items, dict):  # Tümü
            for m, lst in items.items():
                if not lst:
                    continue
                for t in lst:
                    vals.append(f"{m}@{t['id']} — {t.get('title','')}")
        else:
            for t in (items or []):
                vals.append(f"{self.mod_cb.get()}@{t['id']} — {t.get('title','')}")
        self.tpl_cb['values'] = vals
        if vals:
            self.tpl_cb.current(0)

    def _start(self):
        try:
            y = int(self.year.get()); m = int(self.month.get()); assert 1 <= m <= 12
        except Exception:
            messagebox.showerror('Hata', 'Geçerli yıl / ay girin.')
            return
        if not self._guard_api():
            return

        self.log_delete()
        self.stop_event.clear()
        self.btn_start.config(state='disabled')

        t = threading.Thread(target=self._run, args=(y, m), daemon=True)
        self.worker = t
        t.start()

    def _create_and_download(self, mod, start_iso, end_iso, title, tpl_id, folder, log_fn):
        if requests is None:
            raise RuntimeError("requests yok.")
        cfg = REPORT_MODULES[mod]
        payload = {
            "startDate": start_iso.split('T')[0],
            "endDate":   end_iso.split('T')[0],
            "dateCriteria": "DocumentDate",
            "title": title,
            "templateId": tpl_id
        }
        r = requests.post(cfg['create_url'], json=payload, headers=self.headers)
        r.raise_for_status()
        rid = r.json()['id']
        log_fn(f"▶ {mod}: Oluşturuldu (ID {rid})")

        url = cfg['download_url'].format(id=rid)
        dl_hdr = {**self.headers, "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        while not self.stop_event.is_set():
            rr = requests.get(url, headers=dl_hdr)
            if rr.status_code == 200:
                fn = sanitize_filename(rr.headers.get('Content-Disposition', ''), f"{rid}.xlsx")
                fn = re.sub(r'[<>:"/\\|?*]', "_", fn)
                os.makedirs(folder, exist_ok=True)
                full = os.path.join(folder, f"{mod.replace(' ','_')}_{fn}")
                with open(full, 'wb') as f:
                    f.write(rr.content)
                annotate_excel(full, mod, folder.split('_')[-1])

                self.sums[mod] = parse_report(full, mod)
                tax, pay, rej = self.sums[mod]
                log_fn(f"✅ {mod}: Vergi {tax:,.2f} | Ödenecek {pay:,.2f} | Reddedildi {rej:,.2f}")
                break
            elif rr.status_code == 404:
                time.sleep(5)
            else:
                rr.raise_for_status()

    def _run(self, y, m):
        api_name = self.api_cb.get()
        apply_api(self.apis.get(api_name, ''), self.headers)

        ym = f"{y}-{m:02d}"
        today = datetime.now().strftime('%d.%m.%Y')
        self.output_dir = os.path.join(os.getcwd(), f"rapor_{ym}_{today.replace('.', '-')}")

        sd = datetime(y, m, 1)
        ed = datetime(y, m, calendar.monthrange(y, m)[1], 23, 59, 59)
        start_iso = sd.strftime('%Y-%m-%dT00:00:00+03:00')
        end_iso   = ed.strftime('%Y-%m-%dT%H:%M:%S+03:00')

        self._log(f"🔄 Başlıyor: {self.mod_cb.get()} {ym} (API: {api_name})")
        os.makedirs(self.output_dir, exist_ok=True)

        try:
            if self.mod_cb.get() == 'Tümü':
                for sub in REPORT_MODULES.keys():
                    if sub == 'Tümü' or self.stop_event.is_set():
                        continue
                    tpls = list_templates(sub, self.headers) or []
                    if not tpls:
                        continue
                    self._create_and_download(sub, start_iso, end_iso, self.title.get(), tpls[0]['id'],
                                              self.output_dir, self._log)
            else:
                if not self.tpl_cb.get():
                    self._log("⚠️ Şablon seçili değil.")
                else:
                    tpl_id = self.tpl_cb.get().split('@')[-1].split('—')[0].strip()
                    self._create_and_download(self.mod_cb.get(), start_iso, end_iso,
                                              self.title.get(), tpl_id, self.output_dir, self._log)

            if self.sums:
                self._log("\n🔸 Özet:")
                for n, (tx, py, rj) in self.sums.items():
                    self._log(f"  {n:<15} Vergi: {tx:,.2f}  Ödenecek: {py:,.2f}  Reddedildi: {rj:,.2f}")
                out_kdv = (self.sums.get('e-Arşiv', (0,0,0))[0] + self.sums.get('e-Fatura Giden', (0,0,0))[0])
                in_kdv  =  self.sums.get('e-Fatura Gelen', (0,0,0))[0]
                self._log(f"\n💰 KDV Dengesi (Çıkış-Giriş) = {out_kdv:,.2f} − {in_kdv:,.2f} = {out_kdv - in_kdv:,.2f} ₺")

        finally:
            self._log("\n✅ Tüm işlemler tamamlandı." if not self.stop_event.is_set() else "\n🛑 Kullanıcı durdurdu.")
            self.btn_start.config(state='normal')

    def _open_output(self):
        if self.output_dir and os.path.exists(self.output_dir):
            webbrowser.open(self.output_dir)
        else:
            messagebox.showinfo("Bilgi", "Henüz bir çıktı klasörü oluşmadı.")

    # Log helpers
    def _log(self, msg):
        self.log.insert(tk.END, msg + '\n')
        self.log.see(tk.END)

    def log_delete(self):
        self.log.delete('1.0', tk.END)


# ===========================================================================
# Sekme: PDF Eşleştirici
# ===========================================================================

def read_excel_with_dynamic_header(path, keywords=("Seri", "Nokta")):
    """
    İlk 10 satırda başlık değişken olabilir; 'Seri' ve 'Nokta' benzeri başlıkları arar.
    """
    for i in range(10):
        try:
            df = pd.read_excel(path, header=i, engine="openpyxl")
            cols = [str(c).strip().lower() for c in df.columns]
            if all(any(kw.lower() in c for c in cols) for kw in keywords):
                return df
        except Exception:
            continue
    return None


def get_column_name(columns, keyword):
    for col in columns:
        if keyword.lower() in str(col).lower():
            return col
    return None


def find_pdf_containing_serial(serial, pdf_texts: dict):
    s = str(serial).strip()
    if not s:
        return "BOS_SERI"
    for pdf_name, text in pdf_texts.items():
        if s in text:
            return pdf_name
    return "BULUNAMADI"


class PDFMatcherTab(ttk.Frame):
    def __init__(self, parent, apis, headers, invoice_tab_ref: InvoiceTab):
        super().__init__(parent, padding=10)
        self.apis = apis
        self.headers = headers
        self.settings = load_settings()
        self.invoice_tab_ref = invoice_tab_ref

        self.stop_event = threading.Event()
        self.worker = None
        self.last_params = None
        self.output_dir = None

        # UI
        root_cols = 12
        for c in range(root_cols):
            self.columnconfigure(c, weight=1)

        # Üst başlık
        title = ttk.Label(self, text='PDF Eşleştirici', font=('Segoe UI', 12, 'bold'))
        title.grid(row=0, column=0, columnspan=root_cols, sticky='w', pady=(0, 6))

        # API seçimi
        ttk.Label(self, text='API Hesabı (PDF indir için):').grid(row=1, column=0, sticky='w')
        self.api_var = tk.StringVar()
        self.api_cb = ttk.Combobox(self, values=list(apis.keys()), textvariable=self.api_var, state='readonly')
        self.api_cb.grid(row=1, column=1, sticky='ew')
        if apis:
            last_api = self.settings.get('last_api', '')
            if last_api in apis:
                self.api_cb.set(last_api)
            else:
                self.api_cb.current(0)
            apply_api(apis[self.api_cb.get()], self.headers)
        else:
            self.api_cb.set('— API ekleyin (opsiyonel) —')
        self.api_cb.bind('<<ComboboxSelected>>', self._on_api_change)
        
        # Ensure combobox follows theme
        self.api_cb.configure(background=COLORS['card'], foreground=COLORS['text'])

        # Yıl/Ay (e-Fatura Gelen PDF indir)
        now = datetime.now()
        ttk.Label(self, text='Yıl:').grid(row=1, column=2, sticky='e')
        years = [str(now.year + i) for i in range(-10, 15)]
        self.year = ttk.Combobox(self, values=years, width=6, state='readonly')
        self.year.set(str(now.year))
        self.year.grid(row=1, column=3, sticky='w')
        self.year.configure(background=COLORS['card'], foreground=COLORS['text'])

        ttk.Label(self, text='Ay:').grid(row=1, column=4, sticky='e')
        self.month = ttk.Combobox(self, values=[f'{i:02d}' for i in range(1,13)], width=4, state='readonly')
        self.month.set(f'{now.month:02d}')
        self.month.grid(row=1, column=5, sticky='w')
        self.month.configure(background=COLORS['card'], foreground=COLORS['text'])

        self.btn_fetch_incoming = ttk.Button(self, text="📥 e-Fatura Gelen'den PDF indir",
                                             style='Accent.TButton', command=self._fetch_incoming_pdfs)
        self.btn_fetch_incoming.grid(row=1, column=6, sticky='w', padx=6)

        # PDF klasörü & Exceller
        ttk.Label(self, text='PDF Klasörü:').grid(row=2, column=0, sticky='w', pady=(8,0))
        self.pdf_folder_var = tk.StringVar()
        tk.Entry(self, textvariable=self.pdf_folder_var, width=60).grid(row=2, column=1, columnspan=5, sticky='ew', padx=6, pady=(8,0))
        ttk.Button(self, text="📁 Klasör Seç", style='TButton', command=self._browse_pdf_folder).grid(row=2, column=6, sticky='w', pady=(8,0))

        ttk.Label(self, text='Excel Dosyaları:').grid(row=3, column=0, sticky='w')
        self.excel_files_var = tk.StringVar()
        tk.Entry(self, textvariable=self.excel_files_var, width=60).grid(row=3, column=1, columnspan=5, sticky='ew', padx=6)
        ttk.Button(self, text="📄 Dosya Seç", style='TButton', command=self._browse_excel_files).grid(row=3, column=6, sticky='w')

        # Butonlar
        btn_row = ttk.Frame(self)
        btn_row.grid(row=4, column=0, columnspan=root_cols, sticky='e', pady=PADDING['medium'])

        self.btn_start = ttk.Button(btn_row, text="🚀 Başlat", style='Accent.TButton', command=self._start_process)
        self.btn_start.pack(side='left', padx=5)

        self.btn_stop = ttk.Button(btn_row, text="⏹ Durdur", style='Warning.TButton', command=self._stop_process, state='disabled')
        self.btn_stop.pack(side='left', padx=5)

        self.btn_restart = ttk.Button(btn_row, text="🔄 Yeniden Başlat", style='Secondary.TButton', command=self._restart_process, state='disabled')
        self.btn_restart.pack(side='left', padx=5)

        self.btn_open = ttk.Button(btn_row, text="📁 Klasöre Git", style='TButton', command=self._open_output_folder)
        self.btn_open.pack(side='left', padx=5)

        # Log
        log_frame = ttk.Frame(self, style='Card.TFrame')
        log_frame.grid(row=5, column=0, columnspan=root_cols, sticky='nsew', pady=PADDING['medium'])
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(1, weight=1)

        ttk.Label(log_frame, text='Log Kayıtları', font=('Segoe UI', 10, 'bold'),
                  background=COLORS['card'], foreground=COLORS['primary']).grid(row=0, column=0, sticky='w', padx=12, pady=6)
        self.log = scrolledtext.ScrolledText(log_frame, height=18)
        self.log.grid(row=1, column=0, sticky='nsew', padx=12, pady=10)
        style_scrolled_text(self.log)
        try:
            top = self.winfo_toplevel()
            if hasattr(top, 'register_text_widget'):
                top.register_text_widget(self.log)
        except Exception:
            pass

        self.rowconfigure(5, weight=1)

    def _on_api_change(self, _=None):
        api_name = self.api_cb.get()
        if api_name in self.apis:
            apply_api(self.apis[api_name], self.headers)
            self.settings['last_api'] = api_name
            save_settings(self.settings)

    def _guard_api(self) -> bool:
        api_name = self.api_cb.get()
        return api_name in self.apis and bool((self.apis.get(api_name) or "").strip())

    def _browse_pdf_folder(self):
        folder = filedialog.askdirectory(title="PDF klasörünü seçin")
        if folder:
            self.pdf_folder_var.set(folder)

    def _browse_excel_files(self):
        files = filedialog.askopenfilenames(filetypes=[("Excel Dosyaları", "*.xlsx")])
        if files:
            self.excel_files_var.set(";".join(files))

    def _fetch_incoming_pdfs(self):
        if requests is None:
            messagebox.showerror("Eksik Bağımlılık", "requests yüklü değil.\nKomut: pip install requests")
            return
        if not self._guard_api():
            messagebox.showwarning("API Gerekli", "Önce bir API hesabı seçiniz / ekleyiniz.")
            return
        if not self.invoice_tab_ref.save_folder:
            base = filedialog.askdirectory(title="İndirme için ana klasör seçin")
            if not base:
                return
            self.invoice_tab_ref.save_folder = base
            self.invoice_tab_ref.folder_label.config(text=f'Klasör: {base}')
            self.invoice_tab_ref.settings['invoice_folder'] = base
            save_settings(self.invoice_tab_ref.settings)

        try:
            y = int(self.year.get()); m = int(self.month.get())
        except Exception:
            messagebox.showerror("Hata", "Geçerli yıl/ay seçiniz.")
            return

        api_name = self.api_cb.get()
        apply_api(self.apis.get(api_name, ''), self.headers)

        today = datetime.now().strftime('%d.%m.%Y')
        sd = datetime(y, m, 1).strftime('%Y-%m-%dT00:00:00+03:00')
        ed = datetime(y, m, calendar.monthrange(y, m)[1], 23, 59, 59).strftime('%Y-%m-%dT%H:%M:%S+03:00')

        folder_name = f"{api_name} {int(m)}. ay {today} GelenPDF"
        base = os.path.join(self.invoice_tab_ref.save_folder, folder_name)
        os.makedirs(base, exist_ok=True)

        url = BASE_URLS["e-Fatura Gelen"]
        invs = self.invoice_tab_ref._get_invoices(url, sd, ed)
        key = "e-Fatura Gelen"

        for inv in invs:
            num, uid = inv.get('documentNumber') or inv.get('id'), inv.get('id')
            if not num or not uid:
                continue
            d = os.path.join(base, 'pdf'); os.makedirs(d, exist_ok=True)
            r = requests.get(f"{BASE_URLS[key]}/{uid}/pdf", headers=self.headers)
            if r.status_code == 200:
                with open(os.path.join(d, f"{num}.pdf"), 'wb') as f:
                    f.write(r.content)

        self.pdf_folder_var.set(os.path.join(base, 'pdf'))
        self._log(f"✅ e-Fatura Gelen PDF indirildi. Klasör set edildi: {self.pdf_folder_var.get()}")
        webbrowser.open(base)

    def _start_process(self):
        self.stop_event.clear()
        self.log_delete()
        pdf_folder = self.pdf_folder_var.get()
        excel_paths = [p for p in self.excel_files_var.get().split(";") if p.strip()]

        if not pdf_folder or not os.path.isdir(pdf_folder):
            messagebox.showerror("Hata", "Lütfen geçerli bir PDF klasörü seçiniz.")
            return
        if not excel_paths:
            messagebox.showerror("Hata", "Lütfen en az bir Excel dosyası seçiniz.")
            return

        self.output_dir = os.path.join(os.getcwd(), "eslestirme_cikti")
        os.makedirs(self.output_dir, exist_ok=True)

        self.btn_start.config(state='disabled')
        self.btn_stop.config(state='normal')
        self.btn_restart.config(state='disabled')

        self.last_params = (pdf_folder, excel_paths)
        t = threading.Thread(target=self._run_process, args=(pdf_folder, excel_paths), daemon=True)
        self.worker = t
        t.start()

    def _stop_process(self):
        self.stop_event.set()
        self._log("\n❌ İşlem durdurma talebi alındı. Mevcut adım bitince duracak...")
        self.btn_stop.config(state='disabled')

    def _restart_process(self):
        if not self.last_params:
            messagebox.showinfo("Bilgi", "Önce bir işlem başlatın.")
            return
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Bilgi", "Devam eden bir iş var. Önce durdurun.")
            return
        self.stop_event.clear()
        self.log_delete()
        pdf_folder, excel_paths = self.last_params
        self._log("🔁 Yeniden başlatılıyor...")
        self.btn_start.config(state='disabled')
        self.btn_stop.config(state='normal')
        t = threading.Thread(target=self._run_process, args=(pdf_folder, excel_paths), daemon=True)
        self.worker = t
        t.start()

    def _open_output_folder(self):
        if self.output_dir and os.path.exists(self.output_dir):
            webbrowser.open(self.output_dir)
        else:
            messagebox.showwarning("Uyarı", "Çıktı klasörü bulunamadı.")

    def _run_process(self, pdf_folder, excel_paths):
        start_time = time.time()
        self._log(f"\n🔹 PDF klasörü: {pdf_folder}")
        self._log(f"🔹 Excel dosyası sayısı: {len(excel_paths)}")

        pdf_texts = {}
        pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")]
        for idx, pdf_file in enumerate(pdf_files, 1):
            if self.stop_event.is_set():
                self._log("🛑 Durdurma istendi; PDF tarama sonlandırılıyor.")
                return
            try:
                with open(os.path.join(pdf_folder, pdf_file), "rb") as f:
                    reader = PdfReader(f)
                    content = "".join(page.extract_text() or "" for page in reader.pages)
                    pdf_texts[pdf_file] = content
                if idx % 25 == 0:
                    self._log(f"  • {idx}/{len(pdf_files)} PDF okundu...")
            except Exception as e:
                self._log(f"⚠️ {pdf_file} okunamadı: {e}")
        self._log(f"📂 Toplam {len(pdf_files)} PDF dosyası tarandı.\n")

        all_results = {}
        total_rows = 0
        total_matched = 0
        not_found_rows = []

        for excel_path in excel_paths:
            if self.stop_event.is_set():
                self._log("🛑 Durdurma istendi; Excel işleme sonlandırılıyor.")
                return
            t0 = time.time()
            self._log(f"📥 Excel işleniyor: {os.path.basename(excel_path)}")
            df = read_excel_with_dynamic_header(excel_path, keywords=("Seri", "İşlem Noktası"))
            if df is None:
                self._log(f"⚠️ {os.path.basename(excel_path)} içinde 'Seri' veya 'İşlem Noktası' sütunu yok.")
                continue

            seri_col = get_column_name(df.columns, "seri")
            nokta_col = get_column_name(df.columns, "nokta")
            if not seri_col or not nokta_col:
                self._log(f"⚠️ {os.path.basename(excel_path)}: Gerekli sütunlar bulunamadı.")
                continue

            df = df.dropna(subset=[seri_col]).copy()
            total_rows += len(df)
            df["PDF Dosyası"] = df[seri_col].astype(str).apply(lambda x: find_pdf_containing_serial(x, pdf_texts))
            matched_mask = df["PDF Dosyası"].ne("BULUNAMADI") & df["PDF Dosyası"].ne("BOS_SERI")
            total_matched += int(matched_mask.sum())

            not_found_chunk = df.loc[~matched_mask, [seri_col, nokta_col, "PDF Dosyası"]].copy()
            not_found_chunk["Excel"] = os.path.basename(excel_path)
            not_found_rows.append(not_found_chunk)

            for nokta, subdf in df.groupby(nokta_col):
                if nokta not in all_results:
                    all_results[nokta] = []
                all_results[nokta].append(subdf)

            self._log(f"  ✔ {os.path.basename(excel_path)} tamam ({len(df)} satır, eşleşen {int(matched_mask.sum())}). "
                      f"Süre: {time.time()-t0:.2f} sn")

        if not all_results:
            self._log("\n❌ Hiçbir eşleşme bulunamadı.")
            self._finalize_buttons()
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        for nokta, dataframes in all_results.items():
            final_df = pd.concat(dataframes, ignore_index=True)
            filename = f"{str(nokta).strip().replace(' ', '_').lower()}_cikti_{timestamp}.xlsx"
            full_path = os.path.join(self.output_dir, filename)
            final_df.to_excel(full_path, index=False)
            self._log(f"💾 {filename} kaydedildi ({len(final_df)} satır).")

        if not_found_rows:
            nf = pd.concat(not_found_rows, ignore_index=True)
            nf_file = os.path.join(self.output_dir, f"eslesmeyenler_{timestamp}.xlsx")
            nf.to_excel(nf_file, index=False)
            self._log(f"📄 Eşleşmeyenler kaydedildi: {os.path.basename(nf_file)} ({len(nf)} satır).")

        duration = time.time() - start_time
        self._log(f"\n🎉 İşlem tamamlandı. Toplam satır: {total_rows}, Eşleşen: {total_matched}, "
                  f"Eşleşme Oranı: { (100.0*total_matched/total_rows) if total_rows else 0:.2f}%")
        self._log(f"⏱ Toplam süre: {duration:.2f} sn")

        summary_txt = os.path.join(self.output_dir, f"ozet_{timestamp}.txt")
        with open(summary_txt, 'w', encoding='utf-8') as f:
            f.write(f"PDF klasörü: {pdf_folder}\n")
            f.write(f"Excel sayısı: {len(excel_paths)}\n")
            f.write(f"Toplam satır: {total_rows}\nEşleşen: {total_matched}\n"
                    f"Eşleşme Oranı: { (100.0*total_matched/total_rows) if total_rows else 0:.2f}%\n")
            f.write(f"Süre: {duration:.2f} sn\n")
        self._log(f"📝 Özet: {os.path.basename(summary_txt)}")

        self._finalize_buttons()

    def _finalize_buttons(self):
        self.btn_start.config(state='normal')
        self.btn_stop.config(state='disabled')
        self.btn_restart.config(state='normal')

    # Log helpers
    def _log(self, msg):
        self.log.insert(tk.END, msg + '\n')
        self.log.see(tk.END)

    def log_delete(self):
        self.log.delete('1.0', tk.END)


# ========================== BÖLÜM 2/2 – DEVAM ===========================
# ------------------- Envanter Yardımcıları & Sekmesi --------------------

INV_APP_CONFIG = Path.home() / ".ks_envanter_config.json"
TR_MONTH_ABBR = ["Oca","Şub","Mar","Nis","May","Haz","Tem","Ağu","Eyl","Eki","Kas","Ara"]

# ---------- Yardımcılar ----------
def parse_money(x):
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float, np.number)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        m = re.findall(r"\d+", s)
        if not m:
            return 0.0
        joined = "".join(m)
        if len(joined) <= 2:
            return float(joined)
        return float(joined[:-2]+"."+joined[-2:])


def format_tl(amount: float) -> str:
    s = f"{amount:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"₺{s}"


def normalize_text(x):
    return "" if pd.isna(x) else str(x).strip().upper()


def normalize_serial(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if not s or s.casefold() in {"nan", "none", "null"}:
        return ""
    return s.replace(" ", "").upper()


def find_column(df: pd.DataFrame, cands: List[str], default_index: Optional[int]=None) -> str:
    cols = list(df.columns)
    if not cols:
        return "UNKNOWN"
    norm = [str(c).strip().casefold() for c in cols]
    for cand in cands:
        ccf = cand.casefold()
        for i, n in enumerate(norm):
            if ccf in n:
                return cols[i]
    if default_index is not None and 0 <= default_index < len(cols):
        return cols[default_index]
    return cols[-1]


def get_desktop_dir() -> Path:
    try:
        buf = ctypes.create_unicode_buffer(260)
        if ctypes.windll.shell32.SHGetFolderPathW(None, 0x0010, None, 0, buf) == 0:
            p = Path(buf.value)
            if p.exists():
                return p
    except Exception:
        pass
    for c in [Path(os.environ.get("OneDrive",""))/"Masaüstü",
              Path(os.environ.get("OneDrive",""))/"Desktop",
              Path.home()/"Masaüstü", Path.home()/"Desktop"]:
        if c and str(c)!="" and c.exists():
            return c
    return Path.home()


def _col_letter_to_index(s: str) -> Optional[int]:
    s = str(s).strip().upper()
    if not s.isalpha():
        return None
    idx = 0
    for ch in s:
        idx = idx*26 + (ord(ch)-ord('A')+1)
    return idx-1


def _match_override_column(df: pd.DataFrame, ref: str) -> Optional[str]:
    if df is None or df.empty or ref is None:
        return None
    ref = str(ref).strip()
    if ref == "":
        return None
    cols = list(df.columns)
    for c in cols:
        if str(c).strip().casefold() == ref.casefold():
            return c
    for c in cols:
        if ref.casefold() in str(c).strip().casefold():
            return c
    m = re.fullmatch(r"[A-Za-z]+", ref)
    if m:
        pos = _col_letter_to_index(ref)
        if pos is not None and 0 <= pos < len(cols):
            return cols[pos]
    m = re.fullmatch(r"#?(\d+)", ref)
    if m:
        pos = int(m.group(1)) - 1
        if 0 <= pos < len(cols):
            return cols[pos]
    return None


# ---------- Kurallar ----------
DEFAULT_EXCLUSIONS = [
    "KAYIP ZARAR 100","DÜZELTME STOĞU","GETMOBİL REFERANS KOMİSYON",
    "ZARAR","KAMPANYALI SEPET ÜRÜNÜ","KAMPANYALI EKRAN KORUYUCU","TELEFON ZARAR",
]
DEFAULT_BUCKET_RULES: List[Tuple[str,str]] = [
    (r"TELEFON\s*&\s*TABLET\s*&\s*SAAT\s*;\s*YEN[İI]LENM[İI]Ş\s*TELEFON", "YENİLENMİŞ TELEFON"),
    (r"TELEFON\s*&\s*TABLET\s*&\s*SAAT\s*;\s*(YEN[İI]\s*TELEFON|TUŞLU\s*TELEFON)", "YENİ TELEFON"),
    (r"TELEFON\s*&\s*TABLET\s*&\s*SAAT\s*;\s*2\.?\s*EL\s*TELEFON", "2.EL TELEFON"),
    (r"TELEFON\s*&\s*TABLET\s*&\s*SAAT\s*;\s*(2\.?\s*EL\s*TABLET|YEN[İI]\s*TABLET)", "TABLET"),
    (r"TELEFON\s*&\s*TABLET\s*&\s*SAAT\s*;\s*(2\.?\s*EL\s*SAAT|YEN[İI]\s*SAAT)", "SAAT"),
    (r"B[İI]LG[İI]SAYAR\s*&\s*OEM\s*;\s*2\.?\s*EL\s*NOTEBOOK", "NOTEBOOK"),
]
CORE_SERIAL_BUCKETS = [
    "YENİ TELEFON",
    "YENİLENMİŞ TELEFON",
    "2.EL TELEFON",
    "TABLET",
    "SAAT",
    "NOTEBOOK",
]
BUCKET_ORDER = [
    "YENİ TELEFON",
    "YENİLENMİŞ TELEFON",
    "2.EL TELEFON",
    "TABLET",
    "SAAT",
    "NOTEBOOK",
    "SERİ AKSESUAR",
    "AKSESUAR",
]


def apply_bucket_rules(cat_text: str, rules: List[Tuple[str,str]]) -> str:
    txt = normalize_text(cat_text)
    for pattern,bucket in rules:
        if re.search(pattern, txt):
            return bucket
    return "AKSESUAR"


def read_all_sheets(filepath: str, sheet_name_filter: Optional[str]=None) -> pd.DataFrame:
    xls = pd.read_excel(filepath, sheet_name=None, dtype=object)
    frames=[]
    for name, part in xls.items():
        if sheet_name_filter and not re.search(sheet_name_filter, str(name)):
            continue
        df = part.copy()
        df["_kaynak_dosya"]=Path(filepath).name
        df["_kaynak_sayfa"]=str(name)
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def locate_columns(df: pd.DataFrame, col_overrides: Dict[str,str]) -> Tuple[str,str,str,Optional[str],Optional[str]]:
    if df is None or df.empty:
        raise ValueError("Boş veri çerçevesi.")
    def pick(key: str, cands: List[str], default_idx: int):
        ref = (col_overrides.get(key) or "").strip()
        if ref:
            g = _match_override_column(df, ref)
            if g:
                return g
        return find_column(df, cands, default_index=default_idx)
    col_kategori = pick("kategori", ["kategori","kategori adı","ürün grubu","grup"], 2)
    col_adet     = pick("adet",     ["adet","miktar","qty"], 3)
    col_alis     = pick("alis",     ["alış tutarı","alis tutari","kdv dahil alış","kdv dahil alis","[kdv dahil] alış tutarı"], 4)
    prod_ref = (col_overrides.get("ürün adı") or col_overrides.get("urun") or "").strip()
    prod_col=None
    if prod_ref:
        prod_col=_match_override_column(df, prod_ref)
    if not prod_col:
        for cand in ["ürün adı","urun adi","ürün","urun","açıklama","aciklama"]:
            maybe = find_column(df, [cand])
            if maybe in df.columns:
                prod_col=maybe
                break
    serial_col = None
    serial_candidates = [
        "imei",
        "seri",
        "seri no",
        "imei/seri",
        "seri/imei",
        "seri num",
    ]
    cols_norm = [str(c).strip().casefold() for c in df.columns]
    for cand in serial_candidates:
        ccf = cand.casefold()
        for idx, norm_name in enumerate(cols_norm):
            if ccf in norm_name:
                serial_col = df.columns[idx]
                break
        if serial_col:
            break

    return col_kategori,col_adet,col_alis,prod_col,serial_col


def process_frames(dfs: List[pd.DataFrame], col_overrides: Dict[str,str],
                   exclude_keywords: List[str], bucket_rules: List[Tuple[str,str]],
                   drop_totals: bool=True) -> Tuple[pd.DataFrame,pd.DataFrame,pd.DataFrame]:
    if not dfs:
        raise RuntimeError("İşlenecek veri yok.")
    df_all = pd.concat(dfs, ignore_index=True)
    col_kategori,col_adet,col_alis,prod_col,serial_col = locate_columns(df_all, col_overrides)

    out = df_all.copy()
    out["_kategori"] = out[col_kategori].map(normalize_text)
    if serial_col:
        out["_serial"] = out[serial_col].map(normalize_serial)
    else:
        out["_serial"] = ""
    out["_has_serial"] = out["_serial"].str.len() > 0

    if drop_totals:
        out = out[~out["_kategori"].str.contains("TOPLAM", na=False)]
    out = out[out["_kategori"].str.len()>0]

    if exclude_keywords:
        pat = "|".join([re.escape(x) for x in exclude_keywords])
        out = out[~out["_kategori"].str.contains(pat, na=False)]
        if prod_col and prod_col in out.columns:
            prod_norm = out[prod_col].map(normalize_text)
            out = out[~prod_norm.str.contains(pat, na=False)]

    out["_adet"] = pd.to_numeric(out[col_adet].apply(lambda x: str(x).replace(",", ".") if pd.notna(x) else x),
                                 errors="coerce").fillna(0).astype(float)
    out["_alis_kdv_dahil"] = out[col_alis].apply(parse_money).astype(float)
    out["_bucket"] = out["_kategori"].apply(lambda s: apply_bucket_rules(s, bucket_rules))
    core_set = set(CORE_SERIAL_BUCKETS)
    if serial_col:
        serial_accessory_mask = (~out["_bucket"].isin(core_set)) & out["_has_serial"]
        out.loc[serial_accessory_mask, "_bucket"] = "SERİ AKSESUAR"
        out.loc[out["_has_serial"], "_adet"] = 1.0
    accessory_mask = ~out["_bucket"].isin(core_set)
    out.loc[accessory_mask & ~out["_has_serial"], "_bucket"] = "AKSESUAR"

    if serial_col:
        serial_rows = out["_has_serial"] & (out["_serial"].str.len() > 0)
        if bool(serial_rows.any()):
            serial_unique = out.loc[serial_rows].copy()
            serial_unique = serial_unique.drop_duplicates(subset="_serial", keep="first")
            out = pd.concat([out.loc[~serial_rows], serial_unique], ignore_index=True)

    summary = out.groupby("_bucket").agg(ADET=("_adet","sum"), ALIS_CIRO=("_alis_kdv_dahil","sum")).reindex(BUCKET_ORDER).fillna(0.0)

    summary_view = summary.copy()
    summary_view["ADET"] = summary_view["ADET"].round(0).astype(int)
    summary_view["ALIŞ CİRO (KDV DAHİL)"] = summary_view["ALIS_CIRO"].round(2)
    summary_view = summary_view[["ADET","ALIŞ CİRO (KDV DAHİL)"]].reset_index().rename(columns={"_bucket":"KATEGORİ"})

    toplam_adet = int(round(out["_adet"].sum()))
    toplam_ciro = round(float(out["_alis_kdv_dahil"].sum()), 2)
    summary_view.loc[len(summary_view.index)] = ["TÜM ADET / CİRO", toplam_adet, toplam_ciro]

    detail = out.groupby(["_bucket","_kategori"]).agg(ADET=("_adet","sum"), ALIS_CIRO=("_alis_kdv_dahil","sum")).reset_index().sort_values(["_bucket","_kategori"])

    raw_columns: List[Tuple[str,str]] = [
        (col_kategori, "Kategori"),
        ("_adet", "Adet"),
        ("_alis_kdv_dahil", "Alış Tutarı (KDV Dahil)"),
        ("_bucket", "Bucket"),
        ("_kaynak_dosya", "Kaynak Dosya"),
        ("_kaynak_sayfa", "Kaynak Sayfa"),
    ]
    if serial_col:
        raw_columns.append(("_serial", "Seri / IMEI"))
    raw = out[[col for col,_ in raw_columns]].copy()
    raw.columns = [name for _,name in raw_columns]

    return summary_view, detail, raw


# ---------- Excel Export (Kutu tasarım) ----------
def _boxed_title(title_text: str) -> str:
    t = (title_text or "").strip()
    if t:
        return t
    return f"1.{TR_MONTH_ABBR[datetime.now().month-1]}"


def _write_boxed_sheet(writer: pd.ExcelWriter, summary_df: pd.DataFrame, title_text: str=""):
    wb = writer.book
    ws = wb.add_worksheet("Kutu")

    ws.set_column("A:A", 2)
    ws.set_column("B:C", 22)
    ws.set_column("D:E", 22)
    ws.set_column("F:F", 2)

    fmt_header = wb.add_format({"align":"center","valign":"vcenter","bold":True,"border":1,"font_size":12})
    fmt_qty    = wb.add_format({"align":"center","valign":"vcenter","bold":True,"border":1})
    fmt_cat    = wb.add_format({"align":"center","valign":"vcenter","bold":True,"border":1})
    fmt_money  = wb.add_format({"align":"center","valign":"vcenter","bold":True,"border":1})
    fmt_label  = wb.add_format({"align":"center","valign":"vcenter","border":1})
    fmt_spacer = wb.add_format({"border":1})

    row = 0
    ws.merge_range(row, 1, row, 4, _boxed_title(title_text), fmt_header)
    row += 1

    def add_block(r: int, adet: int, kategori: str, ciro: float) -> int:
        ws.merge_range(r,   1, r,   2, f"{int(round(adet))} ADET", fmt_qty)
        ws.merge_range(r,   3, r,   4, kategori,               fmt_cat)
        ws.merge_range(r+1, 1, r+1, 2, format_tl(ciro),        fmt_money)
        ws.merge_range(r+1, 3, r+1, 4, "ALIŞ CİRO",            fmt_label)
        ws.merge_range(r+2, 1, r+2, 2, "", fmt_spacer)
        ws.merge_range(r+2, 3, r+2, 4, "", fmt_spacer)
        return r + 3

    def add_totals(r: int, adet: int, ciro: float) -> int:
        ws.merge_range(r,   1, r,   2, f"{int(round(adet))} ADET", fmt_qty)
        ws.merge_range(r,   3, r,   4, "Tüm Adet",                 fmt_cat)
        ws.merge_range(r+1, 1, r+1, 2, format_tl(ciro),            fmt_money)
        ws.merge_range(r+1, 3, r+1, 4, "Tüm ciro",                 fmt_label)
        return r + 3

    total_row = summary_df.iloc[-1]
    total_adet = int(total_row["ADET"])
    total_ciro = float(total_row["ALIŞ CİRO (KDV DAHİL)"])
    vals = {row["KATEGORİ"]: (int(row["ADET"]), float(row["ALIŞ CİRO (KDV DAHİL)"]))
            for _, row in summary_df.iloc[:-1].iterrows()}

    for cat in BUCKET_ORDER:
        if cat in vals:
            adet, ciro = vals[cat]
            if adet > 0 or ciro > 0:
                row = add_block(row, adet, cat, ciro)

    row = add_totals(row, total_adet, total_ciro)


def export_to_excel(summary_df: pd.DataFrame, detail_df: pd.DataFrame, raw_df: pd.DataFrame,
                    outdir: Path, boxed_title: str="") -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    out_path = outdir / f"ENVANTER_OZET_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
        summary_df.to_excel(writer, sheet_name="Özet", index=False)
        detail_df.to_excel(writer, sheet_name="Detay", index=False)
        raw_df.to_excel(writer, sheet_name="Ham Veriler", index=False)
        _write_boxed_sheet(writer, summary_df, boxed_title)
    return out_path


# ---------- Ayar ----------
def inv_load_config() -> dict:
    if INV_APP_CONFIG.exists():
        try:
            return json.loads(INV_APP_CONFIG.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def inv_save_config(cfg: dict) -> bool:
    try:
        INV_APP_CONFIG.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
        return True
    except Exception:
        return False


# ---------- Envanter Sekmesi ----------
class InventoryTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=10)

        self.inputs: List[str] = []
        self.sheet_filter_var = tk.StringVar(value="")
        self.outdir_var = tk.StringVar(value=str(get_desktop_dir()))

        self.kategori_col = tk.StringVar(value="")
        self.adet_col     = tk.StringVar(value="")
        self.alis_col     = tk.StringVar(value="")
        self.urun_col     = tk.StringVar(value="")

        self.drop_totals_var = tk.BooleanVar(value=True)
        self.bucket_includes: Dict[str, tk.BooleanVar] = {b: tk.BooleanVar(value=True) for b in BUCKET_ORDER}
        self.exclusions: List[str] = list(DEFAULT_EXCLUSIONS)
        self.bucket_rules: List[Tuple[str,str]] = list(DEFAULT_BUCKET_RULES)

        this_month = TR_MONTH_ABBR[datetime.now().month-1]
        self.header_mode  = tk.StringVar(value="free")  # free | month
        self.header_text  = tk.StringVar(value=f"1.{this_month}")
        self.header_month = tk.StringVar(value=this_month)

        self.summary_df: Optional[pd.DataFrame] = None
        self.detail_df : Optional[pd.DataFrame] = None
        self.raw_df    : Optional[pd.DataFrame] = None

        self._build_ui()
        self.apply_config(inv_load_config())

    def _build_ui(self):
        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True)
        self.page_main   = ttk.Frame(nb)
        self.page_mapping= ttk.Frame(nb)
        self.page_rules  = ttk.Frame(nb)
        self.page_preview= ttk.Frame(nb)
        self.page_log    = ttk.Frame(nb)
        nb.add(self.page_main,   text="İşlem")
        nb.add(self.page_mapping,text="Sütun Eşleştirme")
        nb.add(self.page_rules,  text="Bucket & Dışlama")
        nb.add(self.page_preview,text="Önizleme")
        nb.add(self.page_log,    text="Log")

        self._build_main_page()
        self._build_mapping_page()
        self._build_rules_page()
        self._build_preview_page()
        self._build_log_page()

        self.status_var = tk.StringVar(value="Hazır")
        ttk.Label(self, textvariable=self.status_var, anchor="w").pack(fill="x", side="bottom")

    def _build_main_page(self):
        frm = self.page_main
        top = ttk.Frame(frm)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Girdi Dosyaları / Klasör:").pack(side="left")
        ttk.Button(top, text="Excel Ekle", command=self.add_files).pack(side="left", padx=5)
        ttk.Button(top, text="Klasör Ekle", command=self.add_dir).pack(side="left", padx=5)
        ttk.Button(top, text="Temizle", command=self.clear_inputs).pack(side="left", padx=5)

        mid = ttk.Frame(frm)
        mid.pack(fill="both", expand=True, padx=12, pady=6)
        self.lst_inputs = tk.Listbox(mid, height=10)
        self.lst_inputs.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(mid, orient="vertical", command=self.lst_inputs.yview)
        sb.pack(side="left", fill="y")
        self.lst_inputs.config(yscrollcommand=sb.set)
        # Stil ve kayıt
        try:
            style_listbox(self.lst_inputs)
            top = self.winfo_toplevel()
            if hasattr(top, 'register_listbox_widget'):
                top.register_listbox_widget(self.lst_inputs)
        except Exception:
            pass
        right = ttk.Frame(mid)
        right.pack(side="left", fill="y", padx=10)
        ttk.Button(right, text="Seçileni Sil", command=self.remove_selected_input).pack(fill="x", pady=2)
        ttk.Button(right, text="Yukarı", command=lambda: self.move_input(-1)).pack(fill="x", pady=2)
        ttk.Button(right, text="Aşağı", command=lambda: self.move_input(1)).pack(fill="x", pady=2)

        bottom = ttk.LabelFrame(frm, text="Seçenekler")
        bottom.pack(fill="x", padx=12, pady=8)
        ttk.Checkbutton(bottom, text='"TOPLAM" satırlarını hariç tut', variable=self.drop_totals_var).pack(side="left", padx=5)

        sf = ttk.Frame(frm)
        sf.pack(fill="x", padx=12, pady=6)
        ttk.Label(sf, text="Sheet Adı Filtresi (regex):").pack(side="left")
        ttk.Entry(sf, textvariable=self.sheet_filter_var, width=30).pack(side="left", padx=5)

        rf = ttk.LabelFrame(frm, text="Rapor Başlığı (Kutu sayfası)")
        rf.pack(fill="x", padx=12, pady=8)
        ttk.Radiobutton(rf, text="Serbest yaz", variable=self.header_mode, value="free").grid(row=0, column=0, sticky="w", padx=4, pady=2)
        ttk.Radiobutton(rf, text="Ay seç (1.)", variable=self.header_mode, value="month").grid(row=1, column=0, sticky="w", padx=4, pady=2)
        ttk.Entry(rf, textvariable=self.header_text, width=20).grid(row=0, column=1, sticky="w", padx=6)
        ttk.Label(rf, text="Ay:").grid(row=1, column=1, sticky="e")
        ttk.Combobox(rf, textvariable=self.header_month, values=TR_MONTH_ABBR, state="readonly", width=8).grid(row=1, column=2, sticky="w", padx=6)

        of = ttk.Frame(frm)
        of.pack(fill="x", padx=12, pady=6)
        ttk.Label(of, text="Çıktı Klasörü:").pack(side="left")
        ttk.Entry(of, textvariable=self.outdir_var).pack(side="left", fill="x", expand=True, padx=5)
        ttk.Button(of, text="Gözat", command=self.pick_outdir).pack(side="left")

        actions = ttk.Frame(frm)
        actions.pack(fill="x", padx=12, pady=10)
        ttk.Button(actions, text="Önizleme Oluştur", command=self.run_preview).pack(side="left", padx=5)
        ttk.Button(actions, text="Excel'e Aktar", command=self.export_excel).pack(side="left", padx=5)
        ttk.Button(actions, text="Klasörü Aç", command=self.open_outdir).pack(side="left", padx=5)
        ttk.Button(actions, text="Ayarları Kaydet", command=self.save_current_config).pack(side="left", padx=5)
        ttk.Button(actions, text="Ayar Yükle", command=self.load_config_action).pack(side="left", padx=5)

    def _build_mapping_page(self):
        frm = self.page_mapping
        info = ttk.LabelFrame(frm, text="Sütun Eşleştirme (Boş bırakırsan otomatik tespit edilir)")
        info.pack(fill="x", padx=12, pady=10)
        grid = ttk.Frame(info)
        grid.pack(fill="x", padx=8, pady=8)
        ttk.Label(grid, text="Kategori Sütunu:").grid(row=0, column=0, sticky="e")
        ttk.Entry(grid, textvariable=self.kategori_col, width=40).grid(row=0, column=1, sticky="w", padx=6)
        ttk.Label(grid, text="Adet Sütunu:").grid(row=1, column=0, sticky="e")
        ttk.Entry(grid, textvariable=self.adet_col, width=40).grid(row=1, column=1, sticky="w", padx=6)
        ttk.Label(grid, text="[KDV Dahil] Alış Tutarı Sütunu:").grid(row=2, column=0, sticky="e")
        ttk.Entry(grid, textvariable=self.alis_col, width=40).grid(row=2, column=1, sticky="w", padx=6)
        ttk.Label(grid, text="Ürün Adı / Açıklama (opsiyonel):").grid(row=3, column=0, sticky="e")
        ttk.Entry(grid, textvariable=self.urun_col, width=40).grid(row=3, column=1, sticky="w", padx=6)
        ttk.Label(frm, text="İpucu: Başlık adı veya sütun harfi (C/AA) ya da sıra no (3) yazabilirsin.").pack(anchor="w", padx=12)

    def _build_rules_page(self):
        frm = self.page_rules
        inc = ttk.LabelFrame(frm, text="Raporlanacak Bucket'lar")
        inc.pack(fill="x", padx=12, pady=8)
        row = ttk.Frame(inc)
        row.pack(fill="x", padx=6, pady=6)
        for i,b in enumerate(BUCKET_ORDER):
            ttk.Checkbutton(row, text=b, variable=self.bucket_includes[b]).grid(row=0, column=i, sticky="w", padx=8)

        exf = ttk.LabelFrame(frm, text="Hariç Tutulan Anahtarlar")
        exf.pack(fill="both", padx=12, pady=8, expand=True)
        ex_top = ttk.Frame(exf)
        ex_top.pack(fill="x", pady=4)
        self.ex_entry = tk.Entry(ex_top)
        self.ex_entry.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(ex_top, text="Ekle", command=self.add_exclusion).pack(side="left", padx=4)
        ttk.Button(ex_top, text="Sil (Seçili)", command=self.del_exclusion).pack(side="left", padx=4)
        ttk.Button(ex_top, text="Temizle", command=self.clear_exclusions).pack(side="left", padx=4)
        self.ex_list = tk.Listbox(exf, height=6)
        self.ex_list.pack(fill="both", expand=True, padx=6, pady=6)
        # Stil ve kayıt
        try:
            style_listbox(self.ex_list)
            top = self.winfo_toplevel()
            if hasattr(top, 'register_listbox_widget'):
                top.register_listbox_widget(self.ex_list)
        except Exception:
            pass

        brf = ttk.LabelFrame(frm, text="Bucket Kuralları (Regex -> Bucket)")
        brf.pack(fill="both", padx=12, pady=8, expand=True)
        br_top = ttk.Frame(brf)
        br_top.pack(fill="x", pady=4)
        self.rule_pat = tk.Entry(br_top)
        self.rule_pat.pack(side="left", fill="x", expand=True, padx=6)
        self.rule_bucket = ttk.Combobox(br_top, values=BUCKET_ORDER, state="readonly")
        self.rule_bucket.set("YENİ TELEFON")
        self.rule_bucket.pack(side="left", padx=6)
        ttk.Button(br_top, text="Kural Ekle", command=self.add_rule).pack(side="left", padx=4)
        ttk.Button(br_top, text="Kural Sil (Seçili)", command=self.del_rule).pack(side="left", padx=4)

        self.rules_tv = ttk.Treeview(brf, columns=("pattern","bucket"), show="headings", height=6)
        self.rules_tv.heading("pattern", text="Regex Pattern")
        self.rules_tv.heading("bucket", text="Bucket")
        self.rules_tv.pack(fill="both", expand=True, padx=6, pady=6)
        self.refresh_exclusions()
        self.refresh_rules()

    def _build_preview_page(self):
        frm = self.page_preview
        top = ttk.Frame(frm)
        top.pack(fill="x", padx=12, pady=6)
        ttk.Label(top, text="Önizleme tablosu:").pack(side="left")
        self.preview_sel = ttk.Combobox(top, values=["Özet","Detay","Ham Veriler"], state="readonly")
        self.preview_sel.set("Özet")
        self.preview_sel.pack(side="left", padx=8)
        ttk.Button(top, text="Yenile", command=self.refresh_preview).pack(side="left", padx=8)

        self.preview_tv = ttk.Treeview(frm, show="headings")
        self.preview_tv.pack(fill="both", expand=True, padx=12, pady=6)
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.preview_tv.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.preview_tv.xview)
        self.preview_tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

    def _build_log_page(self):
        frm = self.page_log
        self.log = scrolledtext.ScrolledText(frm, height=20, wrap="word")
        self.log.pack(fill="both", expand=True, padx=12, pady=8)
        style_scrolled_text(self.log)
        try:
            top = self.winfo_toplevel()
            if hasattr(top, 'register_text_widget'):
                top.register_text_widget(self.log)
        except Exception:
            pass
        ttk.Button(frm, text="Logu Temizle", command=lambda: self.log.delete(1.0,"end")).pack(padx=12, pady=6, anchor="w")

    # ---- Actions ----
    def add_files(self):
        paths = filedialog.askopenfilenames(title="Excel dosyaları seç", filetypes=[("Excel","*.xlsx *.xls")])
        if not paths:
            return
        for p in paths:
            if p not in self.inputs:
                self.inputs.append(p)
                self.lst_inputs.insert("end", p)
        self.log_write(f"{len(paths)} dosya eklendi.")

    def add_dir(self):
        d = filedialog.askdirectory(title="Klasör seç")
        if not d:
            return
        count=0
        for p in Path(d).glob("*.xlsx"):
            s=str(p)
            if s not in self.inputs:
                self.inputs.append(s)
                self.lst_inputs.insert("end", s)
                count+=1
        self.log_write(f"Klasörden {count} dosya eklendi.")

    def clear_inputs(self):
        self.inputs.clear()
        self.lst_inputs.delete(0,"end")
        self.log_write("Girdi listesi temizlendi.")

    def remove_selected_input(self):
        idx=self.lst_inputs.curselection()
        if not idx:
            return
        i=idx[0]
        self.inputs.pop(i)
        self.lst_inputs.delete(i)
        self.log_write("Seçili dosya kaldırıldı.")

    def move_input(self, delta:int):
        idxs=self.lst_inputs.curselection()
        if not idxs:
            return
        i=idxs[0]
        j=i+delta
        if j<0 or j>=len(self.inputs):
            return
        self.inputs[i],self.inputs[j]=self.inputs[j],self.inputs[i]
        self.lst_inputs.delete(0,"end")
        for s in self.inputs:
            self.lst_inputs.insert("end", s)
        self.lst_inputs.selection_set(j)

    def pick_outdir(self):
        d = filedialog.askdirectory(title="Çıktı klasörü seç")
        if d:
            self.outdir_var.set(d)

    def _collect_overrides(self) -> Dict[str,str]:
        return {
            "kategori": self.kategori_col.get().strip(),
            "adet": self.adet_col.get().strip(),
            "alis": self.alis_col.get().strip(),
            "ürün adı": self.urun_col.get().strip(),
        }

    def run_preview(self):
        if not self.inputs:
            messagebox.showwarning("Uyarı", "Önce Excel dosyaları ekleyin.")
            return
        try:
            dfs=[]
            for path in self.inputs:
                df = read_all_sheets(path, sheet_name_filter=self.sheet_filter_var.get().strip() or None)
                if not df.empty:
                    dfs.append(df)
            if not dfs:
                messagebox.showwarning("Uyarı", "Seçili dosyalarda veri bulunamadı.")
                return

            # İşleme
            summary, detail, raw = process_frames(
                dfs,
                col_overrides=self._collect_overrides(),
                exclude_keywords=self.exclusions,
                bucket_rules=self.bucket_rules,
                drop_totals=self.drop_totals_var.get()
            )

            # Bucket filtreleri uygulanır
            include = {b for b,v in self.bucket_includes.items() if v.get()}
            if include:
                summary_mask = summary["KATEGORİ"].isin(include) | (summary["KATEGORİ"] == "TÜM ADET / CİRO")
                detail_mask = detail["_bucket"].isin(include)
                raw_mask = raw["Bucket"].isin(include)

                if len(summary_mask) != len(summary):
                    raise ValueError("Summary mask length mismatch")
                if len(detail_mask) != len(detail):
                    raise ValueError("Detail mask length mismatch")
                if len(raw_mask) != len(raw):
                    raise ValueError("Raw mask length mismatch")

                summary = summary[summary_mask]
                detail = detail[detail_mask]
                raw = raw[raw_mask]

            self.summary_df = summary
            self.detail_df  = detail.rename(columns={"_bucket":"Bucket","_kategori":"Kategori","ADET":"Adet","ALIS_CIRO":"Alış Ciro"})
            self.raw_df     = raw

            self.refresh_preview()
            self.status_var.set("Önizleme hazır.")
            self.log_write("Önizleme hazır.")
        except Exception as e:
            traceback.print_exc()
            messagebox.showerror("Hata", f"Önizleme oluşturulamadı:\n{e}")

    def export_excel(self):
        if self.summary_df is None or self.detail_df is None or self.raw_df is None:
            messagebox.showinfo("Bilgi", "Önce 'Önizleme Oluştur' deyin.")
            return
        try:
            outdir = Path(self.outdir_var.get().strip() or get_desktop_dir())
            if self.header_mode.get()=="month":
                boxed = f"1.{self.header_month.get().strip() or TR_MONTH_ABBR[datetime.now().month-1]}"
            else:
                boxed = self.header_text.get().strip()
            out_path = export_to_excel(self.summary_df, self.detail_df, self.raw_df, outdir, boxed_title=boxed)
            self.status_var.set(f"Kaydedildi: {out_path.name}")
            self.log_write(f"Excel kaydedildi: {out_path}")
            try:
                if outdir.exists():
                    webbrowser.open(str(outdir))
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Hata", f"Excel'e aktarım başarısız:\n{e}")

    def open_outdir(self):
        try:
            d = Path(self.outdir_var.get().strip())
            if d.exists():
                webbrowser.open(str(d))
            else:
                messagebox.showwarning("Uyarı", "Klasör bulunamadı.")
        except Exception:
            pass

    # --- Exclusions & Rules UI helpers
    def refresh_exclusions(self):
        self.ex_list.delete(0,"end")
        for x in self.exclusions:
            self.ex_list.insert("end", x)

    def add_exclusion(self):
        val = self.ex_entry.get().strip()
        if not val:
            return
        if val not in self.exclusions:
            self.exclusions.append(val)
            self.refresh_exclusions()
            self.ex_entry.delete(0,"end")

    def del_exclusion(self):
        idx = self.ex_list.curselection()
        if not idx:
            return
        i = idx[0]
        try:
            self.exclusions.pop(i)
            self.refresh_exclusions()
        except Exception:
            pass

    def clear_exclusions(self):
        self.exclusions.clear()
        self.refresh_exclusions()

    def refresh_rules(self):
        for i in self.rules_tv.get_children():
            self.rules_tv.delete(i)
        for pat,buck in self.bucket_rules:
            self.rules_tv.insert("", "end", values=(pat,buck))

    def add_rule(self):
        pat = self.rule_pat.get().strip()
        buck = self.rule_bucket.get().strip()
        if not pat or not buck:
            return
        self.bucket_rules.append((pat,buck))
        self.refresh_rules()
        self.rule_pat.delete(0,"end")

    def del_rule(self):
        sel = self.rules_tv.selection()
        if not sel:
            return
        for item in sel:
            vals = self.rules_tv.item(item, "values")
            try:
                self.bucket_rules.remove((vals[0], vals[1]))
            except Exception:
                pass
            self.rules_tv.delete(item)

    def refresh_preview(self):
        df = None
        if self.preview_sel.get()=="Özet":
            df = self.summary_df
        elif self.preview_sel.get()=="Detay":
            df = self.detail_df
        else:
            df = self.raw_df

        for c in self.preview_tv.get_children():
            self.preview_tv.delete(c)
        self.preview_tv["columns"]=()
        if df is None or df.empty:
            return
        cols = list(df.columns)
        self.preview_tv["columns"]=cols
        for c in cols:
            self.preview_tv.heading(c, text=str(c))
            self.preview_tv.column(c, width=max(80, int(9*len(str(c)))))

        # 2000 satır sınırı (UI donmasın)
        for _, row in df.head(2000).iterrows():
            self.preview_tv.insert("", "end", values=[row[c] for c in cols])

    def save_current_config(self):
        cfg = {
            "sheet_filter": self.sheet_filter_var.get(),
            "outdir": self.outdir_var.get(),
            "overrides": self._collect_overrides(),
            "drop_totals": bool(self.drop_totals_var.get()),
            "exclusions": list(self.exclusions),
            "rules": list(self.bucket_rules),
            "includes": {k: bool(v.get()) for k,v in self.bucket_includes.items()},
            "header_mode": self.header_mode.get(),
            "header_text": self.header_text.get(),
            "header_month": self.header_month.get(),
        }
        ok = inv_save_config(cfg)
        self.status_var.set("Ayarlar kaydedildi." if ok else "Ayarlar kaydedilemedi.")
        self.log_write("Ayarlar kaydedildi." if ok else "Ayarlar kaydedilemedi.")

    def load_config_action(self):
        cfg = inv_load_config()
        if not cfg:
            messagebox.showinfo("Bilgi", "Kayıtlı ayar bulunamadı.")
            return
        self.apply_config(cfg)
        self.status_var.set("Ayarlar yüklendi.")
        self.log_write("Ayarlar yüklendi.")

    def apply_config(self, cfg: dict):
        if not cfg:
            return
        self.sheet_filter_var.set(cfg.get("sheet_filter",""))
        self.outdir_var.set(cfg.get("outdir", self.outdir_var.get()))
        ov = cfg.get("overrides", {})
        self.kategori_col.set(ov.get("kategori",""))
        self.adet_col.set(ov.get("adet",""))
        self.alis_col.set(ov.get("alis",""))
        self.urun_col.set(ov.get("ürün adı",""))
        self.drop_totals_var.set(bool(cfg.get("drop_totals", True)))
        self.exclusions = list(cfg.get("exclusions", DEFAULT_EXCLUSIONS))
        self.bucket_rules = [tuple(x) for x in cfg.get("rules", DEFAULT_BUCKET_RULES)]
        incs = cfg.get("includes", {})
        for k,v in self.bucket_includes.items():
            v.set(bool(incs.get(k, True)))
        self.header_mode.set(cfg.get("header_mode","free"))
        self.header_text.set(cfg.get("header_text", self.header_text.get()))
        self.header_month.set(cfg.get("header_month", self.header_month.get()))
        self.refresh_exclusions()
        self.refresh_rules()

    # Log
    def log_write(self, msg: str):
        try:
            self.log.insert("end", msg + "\n")
            self.log.see("end")
        except Exception:
            pass


# ===========================================================================
# HAKKIMIZDA SEKME
# ===========================================================================
class AboutTab(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding=PADDING['xlarge'])
        self.parent = parent
        
        # Ultra-modern header with gradient effect
        header_frame = ttk.Frame(self, style='Elevated.TFrame')
        header_frame.pack(fill='x', padx=PADDING['large'], pady=(PADDING['large'], PADDING['medium']))
        
        # Main content container with better spacing
        content_container = ttk.Frame(header_frame, style='Card.TFrame')
        content_container.pack(fill='both', expand=True, padx=PADDING['medium'], pady=PADDING['medium'])
        
        # Logo with refined positioning (embedded)
        try:
            import io
            import base64
            logo_bytes = base64.b64decode(LOGO_BASE64)
            logo_image = tk.PhotoImage(data=logo_bytes)
            # Slightly larger logo for better visual balance
            self.logo_resized = logo_image.subsample(3, 3)
            logo_label = ttk.Label(content_container, image=self.logo_resized, style='Card.TLabel')
            logo_label.pack(side='left', padx=(0, PADDING['large']))
        except Exception:
            pass

        # Elegant title section with enhanced typography
        title_frame = ttk.Frame(content_container, style='Card.TFrame')
        title_frame.pack(side='left', fill='both', expand=True)
        
        # Main title with refined styling
        main_title = ttk.Label(title_frame, text='Say Muhasebe Envanteri', 
                              style='LargeTitle.TLabel')
        main_title.pack(anchor='w')
        
        # Subtle subtitle with better spacing
        subtitle = ttk.Label(title_frame, 
                            text='Modern fatura yönetimi ve raporlama sistemi • Profesyonel çözüm', 
                            style='Card.TLabel', 
                            font=('Segoe UI', 11),
                            foreground='#64748B')
        subtitle.pack(anchor='w', pady=(6, 0))

        # Refined control buttons with better visual hierarchy
        control_frame = ttk.Frame(content_container, style='Card.TFrame')
        control_frame.pack(side='right', padx=(PADDING['medium'], 0))
        
        # Theme toggle with improved styling
        theme_icon = '🌙' if COLORS == LIGHT_COLORS else '☀️'
        theme_btn = ttk.Button(control_frame, text=theme_icon, width=5,
                              style='Secondary.TButton', command=self.toggle_theme)
        theme_btn.pack(side='top', pady=(0, PADDING['xs']))
        
        # Settings button with consistent styling
        settings_btn = ttk.Button(control_frame, text='⚙️', width=5,
                                 style='Secondary.TButton', command=self.show_settings)
        settings_btn.pack(side='top')

        # Enhanced feature showcase with refined cards
        showcase_frame = ttk.LabelFrame(self, text='✨ Öne Çıkan Özellikler', 
                                       style='TLabelframe', padding=PADDING['large'])
        showcase_frame.pack(fill='both', expand=True, padx=PADDING['large'], pady=PADDING['medium'])
        
        # Improved grid layout with better proportions
        showcase_frame.columnconfigure(0, weight=1)
        showcase_frame.columnconfigure(1, weight=1)
        showcase_frame.rowconfigure(0, weight=1)
        showcase_frame.rowconfigure(1, weight=1)
        
        # Enhanced feature cards with refined content
        features = [
            ('📈', 'Akıllı Raporlama', 'Otomatik KDV hesaplama ve detaylı mali raporlar'),
            ('⚡', 'Hızlı Sorgulama', 'Gelişmiş arama algoritması ile anında sonuçlar'),
            ('🔗', 'PDF Eşleştirme', 'Yapay zeka destekli otomatik belge eşleştirme'),
            ('🎨', 'Premium Tasarım', 'Windows 11 tarzı modern ve şık arayüz')
        ]
        
        for i, (icon, title, desc) in enumerate(features):
            row, col = divmod(i, 2)
            
            # Individual feature card with enhanced styling
            feature_card = ttk.Frame(showcase_frame, style='Elevated.TFrame', padding=PADDING['large'])
            feature_card.grid(row=row, column=col, sticky='nsew', 
                             padx=PADDING['medium'], pady=PADDING['medium'])
            
            # Card header with icon and title
            header_frame = ttk.Frame(feature_card, style='Card.TFrame')
            header_frame.pack(fill='x', pady=(0, PADDING['small']))
            
            ttk.Label(header_frame, text=icon, style='Card.TLabel',
                     font=('Segoe UI', 20)).pack(side='left', padx=(0, PADDING['small']))
            ttk.Label(header_frame, text=title, style='Title.TLabel',
                     font=('Segoe UI', 12, 'bold')).pack(side='left', anchor='w')
            
            # Card description with better formatting
            desc_label = ttk.Label(feature_card, text=desc, style='Card.TLabel',
                                  font=('Segoe UI', 10), wraplength=300,
                                  foreground=COLORS['text_muted'])
            desc_label.pack(anchor='w', fill='x')

        # Enhanced action section with professional CTA
        action_section = ttk.LabelFrame(self, text='🚀 Hızlı Erişim', 
                                      style='TLabelframe', padding=PADDING['large'])
        action_section.pack(fill='x', padx=PADDING['large'], pady=PADDING['medium'])
        
        # Action description
        action_desc = ttk.Label(action_section, 
                              text='Fatura yönetimine hemen başlayın veya mevcut verilerinizi analiz edin.',
                              style='Card.TLabel', font=('Segoe UI', 11),
                              foreground='#64748B')
        action_desc.pack(anchor='w', pady=(0, PADDING['medium']))
        
        # Professional button layout
        button_grid = ttk.Frame(action_section, style='Card.TFrame')
        button_grid.pack(fill='x')
        
        # Primary actions
        primary_frame = ttk.Frame(button_grid, style='Card.TFrame')
        primary_frame.pack(side='left', fill='x', expand=True)
        
        ttk.Button(primary_frame, text='📄  Fatura İşlemleri', 
                  style='Accent.TButton', width=20,
                  command=lambda: self.parent.notebook.select(1)).pack(side='left', padx=(0, PADDING['medium']))
        
        ttk.Button(primary_frame, text='📈  Mali Raporlar', 
                  style='Secondary.TButton', width=18,
                  command=lambda: self.parent.notebook.select(2)).pack(side='left', padx=(0, PADDING['medium']))
        
        # Secondary actions
        secondary_frame = ttk.Frame(button_grid, style='Card.TFrame')
        secondary_frame.pack(side='right')
        
        ttk.Button(secondary_frame, text='🔗  PDF Eşleştir', 
                  style='TButton', width=15,
                  command=lambda: self.parent.notebook.select(3)).pack(side='left', padx=PADDING['small'])

        # Elegant footer with enhanced visual appeal
        footer_section = ttk.Frame(self, style='Card.TFrame')
        footer_section.pack(fill='x', padx=PADDING['large'], pady=PADDING['medium'])
        
        # Separator line
        separator = ttk.Frame(footer_section, height=1, style='Card.TFrame')
        separator.pack(fill='x', pady=(0, PADDING['medium']))
        
        # Footer content with refined messaging
        footer_content = ttk.Frame(footer_section, style='Card.TFrame')
        footer_content.pack(fill='x')
        
        # Version and theme info
        version_text = 'Say Muhasebe v5.0 • Professional Edition'
        theme_hint = f'{theme_icon} Tema: {"Aydınlık" if COLORS == LIGHT_COLORS else "Koyu"} (Değiştirmek için sağ üst köşeye tıklayın)'
        
        ttk.Label(footer_content, text=version_text, style='Card.TLabel',
                 font=('Segoe UI', 10, 'bold')).pack(side='left')
        
        ttk.Label(footer_content, text=theme_hint, style='Card.TLabel',
                 font=('Segoe UI', 9), foreground=COLORS['text_muted']).pack(side='right')
    
    def toggle_theme(self):
        """Toggle between light and dark theme"""
        try:
            main_app = self.winfo_toplevel()
            if hasattr(main_app, '_dark_var'):
                main_app._dark_var.set(not main_app._dark_var.get())
                # Trigger theme change
                main_app._toggle_theme()
        except Exception:
            pass
    
    def show_settings(self):
        """Show settings dialog"""
        try:
            main_app = self.winfo_toplevel()
            if hasattr(main_app, 'show_settings'):
                main_app.show_settings()
            else:
                messagebox.showinfo("⚙️ Ayarlar", "Ayarlar şu anda mevcut değil.")
        except Exception:
            pass


# ===========================================================================
# ANA UYGULAMA
# ===========================================================================
class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("NES Toolkit – Fatura • Rapor • PDF Eşleştirici • Envanter")
        self.settings = load_settings()

        # Tema başlangıcı
        global COLORS
        if self.settings.get('theme', 'dark') == 'dark':
            COLORS = DARK_COLORS.copy()
        else:
            COLORS = LIGHT_COLORS.copy()

        self.style = apply_modern_theme(self)

        # Pencere boyutu
        try:
            self.geometry(self.settings.get("window_size", "1280x860"))
        except Exception:
            self.geometry("1280x860")
        self.minsize(1024, 700)

        # Modern Header Bar
        topbar = ttk.Frame(self, style="Elevated.TFrame", padding=PADDING['medium'])
        topbar.pack(fill="x", padx=PADDING['medium'], pady=PADDING['small'])

        # Modern Logo Section
        logo_frame = ttk.Frame(topbar, style="Card.TFrame")
        logo_frame.pack(side='left')
        
        if HAS_PIL:
            try:
                from PIL import Image, ImageTk
                logo_candidates = ['saylogo3.png', 'sayapp.png']
                for logo_file in logo_candidates:
                    logo_path = Path(__file__).with_name(logo_file)
                    if logo_path.exists():
                        _img = Image.open(str(logo_path)).resize((140, 42), Image.LANCZOS)
                        self._logo_img = ImageTk.PhotoImage(_img)
                        tk.Label(logo_frame, image=self._logo_img, bg=COLORS['card']).pack(side='left', padx=(0, PADDING['medium']))
                        break
            except Exception:
                pass

        # Modern Title Section
        title_frame = ttk.Frame(topbar, style="Card.TFrame")
        title_frame.pack(side='left', padx=(0, PADDING['xlarge']))
        
        ttk.Label(title_frame, text="🚀 NES Toolkit", style="LargeTitle.TLabel").pack(side="left")
        version_label = ttk.Label(title_frame, text="v3.1 Modern", 
                                 background=COLORS['card'], foreground=COLORS['accent'],
                                 font=('Segoe UI', 9, 'bold'))
        version_label.pack(side="left", padx=(PADDING['small'], 0))

        # Modern Control Section (right side)
        controls_frame = ttk.Frame(topbar, style="Card.TFrame")
        controls_frame.pack(side='right')
        
        # Modern theme toggle
        self._dark_var = tk.BooleanVar(value=(self.settings.get('theme','dark')=='dark'))
        def _toggle_theme():
            global COLORS
            self.settings['theme'] = 'dark' if self._dark_var.get() else 'light'
            COLORS = DARK_COLORS.copy() if self._dark_var.get() else LIGHT_COLORS.copy()
            save_settings(self.settings)
            # Reapply theme completely
            self.style = apply_modern_theme(self, self.style)
            # ScrolledText alanlarını da yeniden boya
            for w in getattr(self, '_styled_texts', []):
                style_scrolled_text(w)
            # Listbox'ları yeniden boya
            for lb in getattr(self, '_styled_listboxes', []):
                style_listbox(lb)
            # Update window background
            self.configure(background=COLORS['background'])

        ttk.Button(controls_frame, text="📁  Klasör Aç", style='TButton', 
                  command=lambda: webbrowser.open(os.getcwd())).pack(side="right", padx=PADDING['small'])
        ttk.Checkbutton(controls_frame, text="🌙  Koyu Tema", variable=self._dark_var, 
                       command=_toggle_theme).pack(side='right', padx=PADDING['medium'])

        # ScrolledText kayıt listesi (tema değişince yeniden boyamak için)
        self._styled_texts: List[tk.Text] = []
        def register_text_widget(w):
            self._styled_texts.append(w)
            style_scrolled_text(w)
        self.register_text_widget = register_text_widget

        # Listbox kayıt listesi (tema değişince yeniden boyamak için)
        self._styled_listboxes: List[tk.Listbox] = []
        def register_listbox_widget(w):
            self._styled_listboxes.append(w)
            style_listbox(w)
        self.register_listbox_widget = register_listbox_widget

        # API & Headers
        self.apis = load_apis()
        self.headers = {}

        # Notebook
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True, padx=PADDING['medium'], pady=PADDING['small'])

        # Sekmeler
        self.invoice_tab = InvoiceTab(self.notebook, self.apis, self.headers)
        self.report_tab  = ReportTab(self.notebook, self.apis, self.headers)
        self.pdf_tab     = PDFMatcherTab(self.notebook, self.apis, self.headers, self.invoice_tab)
        self.inv_tab     = InventoryTab(self.notebook)
        self.about_tab   = AboutTab(self.notebook)

        self.notebook.add(self.invoice_tab, text='Fatura')
        self.notebook.add(self.report_tab,  text='Rapor')
        self.notebook.add(self.pdf_tab,     text='PDF Eşleştirici')
        self.notebook.add(self.inv_tab,     text='Envanter')
        self.notebook.add(self.about_tab,   text='Hakkımızda')

        # Son seçili sekme
        try:
            last = int(self.settings.get("last_tab", 0))
            if 0 <= last < len(self.notebook.tabs()):
                self.notebook.select(last)
        except Exception:
            pass

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_tab_changed(self, _evt=None):
        try:
            idx = self.notebook.index(self.notebook.select())
            self.settings["last_tab"] = idx
            save_settings(self.settings)
        except Exception:
            pass

    def _on_close(self):
        try:
            self.settings["window_size"] = self.geometry()
            save_settings(self.settings)
        except Exception:
            pass
        self.destroy()


# ===========================================================================
# main
# ===========================================================================
if __name__ == "__main__":
    try:
        app = MainApp()
        app.mainloop()
    except Exception as e:
        print("Uygulama hata verdi:", e)
        raise
